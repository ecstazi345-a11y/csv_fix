from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

import pandas as pd
import streamlit as st

from services.supabase_client import supabase

try:
    import openpyxl  # noqa: F401

    OPENPYXL_AVAILABLE = True
except Exception:  # noqa: BLE001
    OPENPYXL_AVAILABLE = False

st.set_page_config(layout="wide")

VIEW_DASHBOARD = "monthly_plan_passport_dashboard_v1"
TABLE_PASSPORT_LINES = "monthly_plan_passport_lines"
TABLE_PASSPORTS = "monthly_plan_passports"
TABLE_V2_PLAN_LINES = "monthly_plan_lines_v2"

PRODUCTIVE_HOURS_PER_SHIFT = 8.0

TABLE_LABOR_SUMMARY = "monthly_labor_summary"

PLANNING_MONTH_REFERENCE = [
    "январь-2026",
    "февраль-2026",
    "март-2026",
    "апрель-2026",
    "май-2026",
    "июнь-2026",
    "июль-2026",
    "август-2026",
    "сентябрь-2026",
    "октябрь-2026",
    "ноябрь-2026",
    "декабрь-2026",
]

QUEUE_FILTER_OPTIONS = ["Все", "1 очередь", "2 очередь", "Не определено"]
TITLE_FILTER_OPTIONS = ["Все", "16160-13", "16160-17", "26160-13", "26160-17"]

SCOPE_FIELD_ALIASES: Dict[str, List[str]] = {
    "facility": ["facility_building", "facility", "Титул / объект", "Титул"],
    "discipline": ["construction_discipline", "discipline", "Дисциплина"],
    "system": ["system", "system_label", "systems", "Система"],
}

_SCOPE_EMPTY_TEXT_VALUES = frozenset({"", "nan", "None", "<NA>"})
_INVALID_CREW_LABELS = frozenset({"", "—", "-", "Звено не выбрано"})

NO_PASSPORT_FOR_SLICE_TEXT = (
    "По выбранному срезу утверждённый паспорт месяца ещё не сформирован."
)

PASSPORT_STATUS_RU = {
    "DRAFT": "Черновик",
    "UNDER_REVIEW": "На проверке",
    "APPROVED": "Утверждён",
    "SUPERSEDED": "Заменён",
    "CANCELLED": "Отменён",
}

ADMISSION_STATUS_RU = {
    "APPROVED_TO_EXECUTE": "Допущено",
    "READY_WITH_RISK": "Допущено с риском",
    "APPROVED_BY_OVERRIDE": "Override / риск",
    "BLOCKED": "Заблокировано",
    "WAITING_CHECKS": "Ожидает проверки",
    "NO_CHECKS": "Нет проверок",
}

RISK_ADMISSION_STATUSES = frozenset({"READY_WITH_RISK", "APPROVED_BY_OVERRIDE"})
CLEAN_ADMISSION_STATUSES = frozenset({"APPROVED_TO_EXECUTE"})
DEFERRED_EXCLUDED_STATUSES = frozenset({"BLOCKED"})

FILTER_KEYS = {
    "project": "passport_filter_project",
    "month": "passport_filter_month",
    "queue": "passport_filter_queue",
    "facility": "passport_filter_facility",
    "discipline": "passport_filter_discipline",
    "system": "passport_filter_system",
    "crew": "passport_filter_crew",
    "admission": "passport_filter_admission",
    "risk": "passport_filter_risk",
}

RISK_FILTER_OPTIONS = ["Все", "Только с риском", "Без риска"]


def safe_str(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def safe_num(value: Any) -> float:
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def money_ru(value: Any) -> str:
    amount = safe_num(value)
    if amount == 0:
        return "—"
    return f"{amount:,.0f} ₽".replace(",", " ")


def pct_display(value: float) -> str:
    if value <= 0:
        return "—"
    return f"{value:.1f}".replace(".", ",") + " %"


def qty_display(value: float) -> str:
    if value <= 0:
        return "—"
    return f"{value:,.3f}".replace(",", " ")


def hours_display(value: float) -> str:
    if value <= 0:
        return "—"
    return f"{value:,.1f}".replace(",", " ")


def kpi_hours_display(value: float) -> str:
    if value <= 0:
        return "—"
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}".replace(",", " ") + " чел·ч"
    return f"{value:,.1f}".replace(",", " ").replace(".", ",") + " чел·ч"


def kpi_people_display(value: float) -> str:
    count = int(round(value))
    if count <= 0:
        return "—"
    return f"{count} чел."


def kpi_days_display(value: float) -> str:
    if value <= 0:
        return "—"
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}".replace(",", " ") + " дн."
    return f"{value:,.1f}".replace(",", " ").replace(".", ",") + " дн."


def kpi_norm_display(value: float) -> str:
    if value <= 0:
        return "—"
    text = f"{value:,.2f}".replace(",", " ").replace(".", ",")
    if "," in text:
        text = text.rstrip("0").rstrip(",")
    return text + " чел·ч / ед."


def kpi_shift_output_display(value: float) -> str:
    if value <= 0:
        return "—"
    if abs(value - round(value)) < 1e-9:
        return f"{int(round(value)):,}".replace(",", " ") + " ед./смен"
    text = f"{value:,.3f}".replace(",", " ").replace(".", ",")
    return text.rstrip("0").rstrip(",") + " ед./смен"


def compute_duration_shifts(hours: float, crew_size: float) -> float:
    safe_hours = safe_num(hours)
    if safe_hours <= 0:
        return 0.0
    safe_crew = max(int(safe_num(crew_size)), 1) if safe_num(crew_size) > 0 else 1
    return safe_hours / (safe_crew * PRODUCTIVE_HOURS_PER_SHIFT)


def filter_options(df: pd.DataFrame, col: str) -> List[str]:
    if df.empty or col not in df.columns:
        return ["Все"]
    vals = df[col].dropna().astype(str).str.strip()
    vals = vals[vals != ""].unique().tolist()
    return ["Все"] + sorted(vals)


def merge_filter_options(base: List[str], data_values: List[str]) -> List[str]:
    seen: set[str] = set()
    merged: List[str] = ["Все"]
    for value in base:
        if value == "Все" or value in seen:
            continue
        seen.add(value)
        merged.append(value)
    for value in sorted({safe_str(v) for v in data_values if safe_str(v)}):
        if value in seen:
            continue
        seen.add(value)
        merged.append(value)
    return merged


def reference_filter_options(ref_df: pd.DataFrame, column: str) -> List[str]:
    if ref_df.empty or column not in ref_df.columns:
        return ["Все"]
    values = ref_df[column].dropna().astype(str).str.strip()
    unique = sorted({value for value in values if value and value != "—"})
    return ["Все", *unique] if unique else ["Все"]


def sync_selectbox_option(key: str, options: List[str]) -> None:
    if not options:
        return
    current = st.session_state.get(key)
    if current not in options:
        st.session_state[key] = options[0]


def _is_bare_bhk_project_label(value: Any) -> bool:
    text = safe_str(value).upper()
    return text in {"БХК", "BHK"}


def _is_valid_project_filter_option(value: Any) -> bool:
    text = safe_str(value)
    if not text:
        return False
    return not _is_bare_bhk_project_label(text)


def normalize_scope_project_codes(df: pd.DataFrame) -> pd.Series:
    if df.empty or "project_code" not in df.columns:
        return pd.Series([""] * len(df), index=df.index)
    result = df["project_code"].fillna("").astype(str).str.strip()
    bare_bhk = result.apply(_is_bare_bhk_project_label)
    return result.mask(bare_bhk, "")


def scope_pick_series(df: pd.DataFrame, field: str, default: str = "") -> pd.Series:
    aliases = SCOPE_FIELD_ALIASES.get(field, [field])
    present = [col for col in aliases if col in df.columns]
    if not present:
        return pd.Series([default] * len(df), index=df.index)
    result = pd.Series(pd.NA, index=df.index, dtype=object)
    for col in present:
        candidate = df[col].fillna("").astype(str).str.strip()
        candidate = candidate.mask(candidate.isin(_SCOPE_EMPTY_TEXT_VALUES), pd.NA)
        result = result.fillna(candidate)
    return result.fillna(default).astype(str).str.strip()


@st.cache_data(ttl=300)
def load_planning_scope_filter_reference() -> pd.DataFrame:
    try:
        from services.monthly_scope_adjustments import load_scope

        raw = load_scope()
    except Exception:  # noqa: BLE001
        raw = pd.DataFrame()
    if raw.empty:
        return pd.DataFrame(
            columns=["project_code", "construction_queue", "facility", "discipline", "system"]
        )
    out = pd.DataFrame(
        {
            "project_code": normalize_scope_project_codes(raw),
            "facility": scope_pick_series(raw, "facility"),
            "discipline": scope_pick_series(raw, "discipline"),
            "system": scope_pick_series(raw, "system"),
        }
    )
    out["construction_queue"] = out["facility"].apply(derive_construction_queue)
    return out


@st.cache_data(ttl=300)
def load_labor_summary_crew_codes() -> List[str]:
    try:
        response = (
            supabase.table(TABLE_LABOR_SUMMARY)
            .select("crew_code")
            .limit(5000)
            .execute()
        )
        df = pd.DataFrame(response.data or [])
        if df.empty or "crew_code" not in df.columns:
            return []
        vals = df["crew_code"].dropna().astype(str).str.strip()
        return sorted({value for value in vals if value and value not in _INVALID_CREW_LABELS})
    except Exception:  # noqa: BLE001
        return []


def planning_month_filter_options(v2_df: pd.DataFrame) -> List[str]:
    months = list(PLANNING_MONTH_REFERENCE)
    seen = set(months)
    if not v2_df.empty and "month_key" in v2_df.columns:
        for value in v2_df["month_key"].dropna().astype(str).str.strip():
            if not value or value == "Все месяца" or value in seen:
                continue
            seen.add(value)
            months.append(value)
    return ["Все", *months]


def planning_project_filter_options(scope_ref: pd.DataFrame) -> List[str]:
    if scope_ref.empty or "project_code" not in scope_ref.columns:
        return ["Все"]
    values = scope_ref["project_code"].dropna().astype(str).str.strip()
    unique = sorted({value for value in values if _is_valid_project_filter_option(value)})
    return ["Все", *unique] if unique else ["Все"]


def planning_title_filter_options(scope_ref: pd.DataFrame) -> List[str]:
    data_values: List[str] = []
    if not scope_ref.empty and "facility" in scope_ref.columns:
        data_values = scope_ref["facility"].dropna().astype(str).tolist()
    return merge_filter_options(TITLE_FILTER_OPTIONS, data_values)


def planning_discipline_filter_options(scope_ref: pd.DataFrame) -> List[str]:
    return reference_filter_options(scope_ref, "discipline")


def planning_system_filter_options(scope_ref: pd.DataFrame, v2_df: pd.DataFrame) -> List[str]:
    options = reference_filter_options(scope_ref, "system")
    extra: List[str] = []
    if not v2_df.empty and "system" in v2_df.columns:
        extra = v2_df["system"].dropna().astype(str).tolist()
    if len(options) <= 1 and not extra:
        return ["Все"]
    return merge_filter_options(options, extra)


def planning_crew_filter_options(
    v2_df: pd.DataFrame,
    passport_df: pd.DataFrame,
) -> List[str]:
    values: List[str] = list(load_labor_summary_crew_codes())
    if not v2_df.empty:
        for col in ("crew", "crew_code"):
            if col in v2_df.columns:
                values.extend(v2_df[col].dropna().astype(str).tolist())
    if not passport_df.empty and "crew_label" in passport_df.columns:
        values.extend(passport_df["crew_label"].dropna().astype(str).tolist())
    unique = sorted(
        {
            safe_str(value)
            for value in values
            if safe_str(value) and safe_str(value) not in _INVALID_CREW_LABELS
        }
    )
    return ["Все", *unique] if unique else ["Все"]


def derive_construction_queue(title: Any) -> str:
    text = safe_str(title)
    if "16160-13" in text or "16160-17" in text:
        return "1 очередь"
    if "26160-13" in text or "26160-17" in text:
        return "2 очередь"
    return "Не определено"


def title_matches_filter(title_value: Any, selected_title: str) -> bool:
    text = safe_str(title_value)
    if not text:
        return False
    return selected_title in text or text == selected_title


def is_risk_row(row: pd.Series) -> bool:
    status = safe_str(row.get("admission_status"))
    if status in RISK_ADMISSION_STATUSES:
        return True
    return bool(row.get("management_override"))


def admission_status_label(value: Any) -> str:
    raw = safe_str(value)
    if not raw:
        return "—"
    return ADMISSION_STATUS_RU.get(raw, raw)


def risk_reason_text(row: pd.Series) -> str:
    for field in ("override_reason", "override_risk_comment", "override_basis"):
        text = safe_str(row.get(field))
        if text:
            return text
    return "—"


@st.cache_data(ttl=300)
def load_v2_plan_lines() -> pd.DataFrame:
    try:
        response = supabase.table(TABLE_V2_PLAN_LINES).select("*").limit(10000).execute()
        return pd.DataFrame(response.data or [])
    except Exception:  # noqa: BLE001
        return pd.DataFrame()


@st.cache_data(ttl=300)
def load_passport_dataset() -> pd.DataFrame:
    try:
        lines_resp = (
            supabase.table(TABLE_PASSPORT_LINES).select("*").limit(10000).execute()
        )
        lines = pd.DataFrame(lines_resp.data or [])
        if lines.empty:
            view_resp = supabase.table(VIEW_DASHBOARD).select("*").limit(10000).execute()
            return pd.DataFrame(view_resp.data or [])

        headers_resp = (
            supabase.table(TABLE_PASSPORTS)
            .select(
                "passport_id, passport_status, passport_name, approved_by, approved_at"
            )
            .limit(1000)
            .execute()
        )
        headers = pd.DataFrame(headers_resp.data or [])
        if not headers.empty:
            lines = lines.merge(headers, on="passport_id", how="left")
        return lines
    except Exception:  # noqa: BLE001
        try:
            view_resp = supabase.table(VIEW_DASHBOARD).select("*").limit(10000).execute()
            return pd.DataFrame(view_resp.data or [])
        except Exception:  # noqa: BLE001
            return pd.DataFrame()


def enrich_passport_dataframe(df: pd.DataFrame, v2_df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()

    v2_lookup: Dict[str, Dict[str, Any]] = {}
    if not v2_df.empty and "plan_line_id" in v2_df.columns:
        for _, v2_row in v2_df.iterrows():
            pid = safe_str(v2_row.get("plan_line_id"))
            if pid:
                v2_lookup[pid] = v2_row.to_dict()

    systems: List[str] = []
    iwps: List[str] = []
    crew_sizes: List[float] = []
    crew_labels: List[str] = []

    for _, row in out.iterrows():
        line_id = safe_str(row.get("line_id"))
        v2_row = v2_lookup.get(line_id, {})
        system = safe_str(v2_row.get("system")) or "—"
        iwp = safe_str(v2_row.get("iwp")) or "—"
        crew_size_raw = safe_num(v2_row.get("crew_size"))
        crew_id = safe_str(row.get("crew_id")) or safe_str(v2_row.get("crew")) or "—"
        systems.append(system if system else "—")
        iwps.append(iwp if iwp else "—")
        crew_sizes.append(crew_size_raw)
        crew_labels.append(crew_id if crew_id else "—")

    out["system_label"] = systems
    out["iwp"] = iwps
    out["crew_size"] = crew_sizes
    out["crew_label"] = crew_labels

    if "plan_value" in out.columns:
        out["plan_value_num"] = out["plan_value"].apply(safe_num)
    else:
        out["plan_value_num"] = 0.0
    if "planned_qty" in out.columns:
        out["planned_qty_num"] = out["planned_qty"].apply(safe_num)
    else:
        out["planned_qty_num"] = 0.0
    if "required_hours" in out.columns:
        out["required_hours_num"] = out["required_hours"].apply(safe_num)
    else:
        out["required_hours_num"] = 0.0
    if "labor_cost" in out.columns:
        out["labor_cost_num"] = out["labor_cost"].apply(safe_num)
    else:
        out["labor_cost_num"] = 0.0

    durations: List[str] = []
    labor_pcts: List[str] = []
    for _, row in out.iterrows():
        hours = safe_num(row.get("required_hours_num"))
        crew_size = safe_num(row.get("crew_size"))
        plan_val = safe_num(row.get("plan_value_num"))
        labor_cost = safe_num(row.get("labor_cost_num"))
        if hours > 0 and crew_size > 0:
            shifts = hours / (crew_size * PRODUCTIVE_HOURS_PER_SHIFT)
            durations.append(f"{shifts:,.1f}".replace(",", " ").replace(".", ",") + " смен")
        else:
            durations.append("—")
        if plan_val > 0 and labor_cost > 0:
            labor_pcts.append(pct_display(labor_cost / plan_val * 100.0))
        else:
            labor_pcts.append("—")

    out["duration_shifts_display"] = durations
    out["labor_to_plan_pct_display"] = labor_pcts
    if "facility_building" in out.columns:
        out["title_display"] = out["facility_building"].apply(lambda v: safe_str(v) or "—")
    else:
        out["title_display"] = "—"
    out["queue_display"] = out["title_display"].apply(derive_construction_queue)
    if "construction_discipline" in out.columns:
        out["discipline_display"] = out["construction_discipline"].apply(
            lambda v: safe_str(v) or "—"
        )
    else:
        out["discipline_display"] = "—"
    if "admission_status" in out.columns:
        out["admission_status_label"] = out["admission_status"].apply(admission_status_label)
    else:
        out["admission_status_label"] = "—"
    out["is_risk"] = out.apply(is_risk_row, axis=1)
    out["risk_reason_display"] = out.apply(risk_reason_text, axis=1)
    return out


def apply_passport_filters(
    df: pd.DataFrame,
    *,
    project: str,
    month: str,
    queue: str,
    facility: str,
    discipline: str,
    system: str,
    crew: str,
    admission: str,
    risk: str,
) -> pd.DataFrame:
    if df.empty:
        return df
    result = df.copy()
    if project != "Все" and "project_code" in result.columns:
        result = result[result["project_code"].astype(str) == project]
    if month != "Все" and "month_key" in result.columns:
        result = result[result["month_key"].astype(str) == month]
    if queue != "Все" and "queue_display" in result.columns:
        result = result[result["queue_display"].astype(str) == queue]
    if facility != "Все" and "title_display" in result.columns:
        result = result[
            result["title_display"].apply(lambda v: title_matches_filter(v, facility))
        ]
    if discipline != "Все" and "discipline_display" in result.columns:
        result = result[result["discipline_display"].astype(str) == discipline]
    if system != "Все" and "system_label" in result.columns:
        result = result[result["system_label"].astype(str) == system]
    if crew != "Все" and "crew_label" in result.columns:
        result = result[result["crew_label"].astype(str) == crew]
    if admission != "Все" and "admission_status_label" in result.columns:
        result = result[result["admission_status_label"].astype(str) == admission]
    if risk == "Только с риском" and "is_risk" in result.columns:
        result = result[result["is_risk"].astype(bool)]
    elif risk == "Без риска" and "is_risk" in result.columns:
        result = result[~result["is_risk"].astype(bool)]
    return result


def compute_passport_summary(df: pd.DataFrame) -> Dict[str, Any]:
    if df.empty:
        return {}
    total_value = float(df["plan_value_num"].sum())
    total_hours = float(df["required_hours_num"].sum())
    total_labor = float(df["labor_cost_num"].sum())
    risk_df = df[df["is_risk"].astype(bool)] if "is_risk" in df.columns else pd.DataFrame()
    if "admission_status" in df.columns:
        admitted_df = df[df["admission_status"].astype(str).isin(CLEAN_ADMISSION_STATUSES)]
        deferred_df = df[df["admission_status"].astype(str).isin(DEFERRED_EXCLUDED_STATUSES)]
    else:
        admitted_df = df[~df["is_risk"].astype(bool)] if "is_risk" in df.columns else df
        deferred_df = pd.DataFrame()
    boq_count = df["boq_code"].nunique() if "boq_code" in df.columns else len(df)
    row_count = len(df)
    if "crew_label" in df.columns:
        labels = df["crew_label"].astype(str).str.strip()
        crew_count = int(labels[~labels.isin(_INVALID_CREW_LABELS)].nunique())
    else:
        crew_count = 0
    risk_value = float(risk_df["plan_value_num"].sum()) if not risk_df.empty else 0.0
    admitted_value = float(admitted_df["plan_value_num"].sum()) if not admitted_df.empty else 0.0
    deferred_excluded_value = (
        float(deferred_df["plan_value_num"].sum()) if not deferred_df.empty else 0.0
    )
    avg_boq_value = total_value / boq_count if boq_count else 0.0
    risk_share = risk_value / total_value * 100.0 if total_value > 0 else 0.0
    admitted_share = admitted_value / total_value * 100.0 if total_value > 0 else 0.0
    labor_pct = total_labor / total_value * 100.0 if total_value > 0 else 0.0

    duration_shifts_values: List[float] = []
    norm_hours_values: List[float] = []
    shift_output_values: List[float] = []
    for _, row in df.iterrows():
        hours = safe_num(row.get("required_hours_num"))
        crew_size = safe_num(row.get("crew_size"))
        qty = safe_num(row.get("planned_qty_num"))
        duration = compute_duration_shifts(hours, crew_size)
        if duration > 0:
            duration_shifts_values.append(duration)
            if qty > 0:
                shift_output_values.append(qty / duration)
        if qty > 0 and hours > 0:
            norm_hours_values.append(hours / qty)

    required_people = 0
    if "crew_label" in df.columns and "crew_size" in df.columns:
        crew_people = (
            df.groupby("crew_label")["crew_size"]
            .apply(lambda sizes: max(safe_num(value) for value in sizes) if len(sizes) else 0.0)
        )
        required_people = int(
            sum(
                safe_num(value)
                for label, value in crew_people.items()
                if safe_str(label) and safe_str(label) not in _INVALID_CREW_LABELS
            )
        )

    return {
        "row_count": row_count,
        "boq_count": boq_count,
        "planned_qty": float(df["planned_qty_num"].sum()),
        "clean_count": len(admitted_df),
        "risk_count": len(risk_df),
        "deferred_excluded_count": len(deferred_df),
        "total_value": total_value,
        "risk_value": risk_value,
        "clean_value": admitted_value,
        "deferred_excluded_value": deferred_excluded_value,
        "avg_boq_value": avg_boq_value,
        "risk_share_pct": risk_share,
        "admitted_share_pct": admitted_share,
        "total_hours": total_hours,
        "total_labor": total_labor,
        "crew_count": crew_count,
        "required_people": required_people,
        "avg_duration_days": (
            sum(duration_shifts_values) / len(duration_shifts_values)
            if duration_shifts_values
            else 0.0
        ),
        "avg_norm_hours": (
            sum(norm_hours_values) / len(norm_hours_values) if norm_hours_values else 0.0
        ),
        "avg_shift_output": (
            sum(shift_output_values) / len(shift_output_values) if shift_output_values else 0.0
        ),
        "labor_pct": labor_pct,
    }


def build_breakdown_table(df: pd.DataFrame, group_col: str, label_col: str) -> pd.DataFrame:
    if df.empty or group_col not in df.columns:
        return pd.DataFrame()
    total_value = float(df["plan_value_num"].sum())
    grouped = (
        df.groupby(group_col, dropna=False)
        .agg(
            rows=("boq_code", "count"),
            plan_value=("plan_value_num", "sum"),
            labor_hours=("required_hours_num", "sum"),
            labor_cost=("labor_cost_num", "sum"),
        )
        .reset_index()
    )
    grouped["share_pct"] = grouped["plan_value"].apply(
        lambda v: (safe_num(v) / total_value * 100.0) if total_value > 0 else 0.0
    )
    grouped = grouped.rename(
        columns={
            group_col: label_col,
            "rows": "Строк",
            "plan_value": "Стоимость работ",
            "labor_hours": "Трудозатраты, чел·ч",
            "labor_cost": "Стоимость труда",
            "share_pct": "Доля, %",
        }
    )
    grouped["Стоимость работ"] = grouped["Стоимость работ"].apply(money_ru)
    grouped["Трудозатраты, чел·ч"] = grouped["Трудозатраты, чел·ч"].apply(hours_display)
    grouped["Стоимость труда"] = grouped["Стоимость труда"].apply(money_ru)
    grouped["Доля, %"] = grouped["Доля, %"].apply(pct_display)
    return grouped


def build_passport_display_table(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    rows: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        crew_size = safe_num(row.get("crew_size"))
        rows.append(
            {
                "Проект": safe_str(row.get("project_code")) or "—",
                "Месяц": safe_str(row.get("month_key")) or "—",
                "Титул": safe_str(row.get("title_display")) or "—",
                "Дисциплина": safe_str(row.get("discipline_display")) or "—",
                "Система": safe_str(row.get("system_label")) or "—",
                "IWP": safe_str(row.get("iwp")) or "—",
                "BOQ-код": safe_str(row.get("boq_code")) or "—",
                "Наименование работ": safe_str(row.get("boq_name")) or "—",
                "Ед.": safe_str(row.get("unit_of_measure")) or "—",
                "Плановый объём": qty_display(safe_num(row.get("planned_qty_num"))),
                "Стоимость работ": money_ru(row.get("plan_value_num")),
                "Звено": safe_str(row.get("crew_label")) or "—",
                "Людей в звене": str(int(crew_size)) if crew_size > 0 else "—",
                "Трудозатраты, чел·ч": hours_display(safe_num(row.get("required_hours_num"))),
                "Длительность, смен": safe_str(row.get("duration_shifts_display")) or "—",
                "Стоимость труда": money_ru(row.get("labor_cost_num")),
                "Труд / стоимость работ, %": safe_str(row.get("labor_to_plan_pct_display")) or "—",
                "Статус допуска": safe_str(row.get("admission_status_label")) or "—",
                "Риск / причина риска": safe_str(row.get("risk_reason_display")) or "—",
                "passport_id": safe_str(row.get("passport_id")) or "—",
            }
        )
    return pd.DataFrame(rows)


def inject_passport_page_styles() -> None:
    st.markdown(
        """
        <style>
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.passport-filter-heading) {
            padding-top: 0.35rem;
        }
        .passport-filter-heading {
            margin: 0 0 0.35rem 0;
            font-size: 1.05rem;
            font-weight: 600;
            color: #0f172a;
        }
        .passport-status-compact {
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #f8fafc;
            padding: 0.65rem 0.85rem;
            margin: 0.35rem 0 0.85rem 0;
        }
        .passport-status-compact-title {
            font-size: 0.72rem;
            font-weight: 700;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin: 0 0 0.45rem 0;
        }
        .passport-status-grid {
            display: grid;
            grid-template-columns: repeat(5, minmax(0, 1fr));
            gap: 0.65rem;
        }
        .passport-status-item-label {
            font-size: 0.68rem;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.04em;
            margin-bottom: 0.1rem;
        }
        .passport-status-item-value {
            font-size: 0.92rem;
            font-weight: 600;
            color: #0f172a;
            line-height: 1.25;
            word-break: break-word;
        }
        .v2-kpi-row {
            display: grid;
            grid-template-columns: repeat(5, minmax(0, 1fr));
            gap: 0.75rem;
            margin: 0.45rem 0 0.15rem 0;
        }
        .v2-kpi-row--7 {
            grid-template-columns: repeat(7, minmax(0, 1fr));
            gap: 0.55rem;
        }
        .v2-kpi-card {
            display: flex;
            align-items: flex-start;
            gap: 0.65rem;
            padding: 0.85rem 0.95rem;
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #ffffff;
            min-height: 78px;
        }
        .v2-kpi-card-icon {
            flex: 0 0 34px;
            width: 34px;
            height: 34px;
            border-radius: 8px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 0.9rem;
            font-weight: 700;
        }
        .v2-kpi-card--total .v2-kpi-card-icon { background: #E8EEF5; color: #475F7B; }
        .v2-kpi-card--ready .v2-kpi-card-icon { background: #E7F5EE; color: #2F6B4F; }
        .v2-kpi-card--risk .v2-kpi-card-icon { background: #F9EDE8; color: #A65F45; }
        .v2-kpi-card--open .v2-kpi-card-icon { background: #E6EEF8; color: #2E5B9A; }
        .v2-kpi-card--muted .v2-kpi-card-icon { background: #f1f5f9; color: #64748b; }
        .v2-kpi-card-label {
            font-size: 0.72rem;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.04em;
            margin-bottom: 0.15rem;
        }
        .v2-kpi-card-value {
            font-size: 1.25rem;
            font-weight: 700;
            color: #0f172a;
            line-height: 1.15;
        }
        .passport-kpi-panel {
            border: 1px solid #e2e8f0;
            border-radius: 10px;
            background: #f8fafc;
            padding: 0.75rem 0.85rem 0.65rem;
            margin: 0.35rem 0 0.85rem 0;
        }
        .passport-kpi-group + .passport-kpi-group {
            margin-top: 0.65rem;
            padding-top: 0.65rem;
            border-top: 1px solid #e2e8f0;
        }
        .passport-kpi-group-title {
            font-size: 0.7rem;
            font-weight: 700;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin: 0 0 0.45rem 0;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def _kpi_card_html(label: str, value: str, variant: str) -> str:
    icons = {"total": "∑", "ready": "✓", "risk": "!", "open": "○", "muted": "·"}
    icon = icons.get(variant, "·")
    return (
        f'<div class="v2-kpi-card v2-kpi-card--{variant}">'
        f'<div class="v2-kpi-card-icon">{icon}</div>'
        f"<div>"
        f'<div class="v2-kpi-card-label">{label}</div>'
        f'<div class="v2-kpi-card-value">{value}</div>'
        f"</div></div>"
    )


def render_passport_summary_panel(summary: Dict[str, Any]) -> None:
    if not summary:
        st.info("Нет данных для сводки.")
        return

    deferred_count = summary.get("deferred_excluded_count", 0)
    deferred_value = summary.get("deferred_excluded_value", 0.0)

    volume_cards = "".join(
        [
            _kpi_card_html("Строк в паспорте", str(summary["row_count"]), "total"),
            _kpi_card_html("BOQ-кодов", str(summary["boq_count"]), "total"),
            _kpi_card_html("Плановый объём", qty_display(summary["planned_qty"]), "total"),
            _kpi_card_html("Допущено", str(summary["clean_count"]), "ready"),
            _kpi_card_html("С риском", str(summary["risk_count"]), "risk"),
            _kpi_card_html(
                "Отложено / исключено",
                str(deferred_count) if deferred_count else "—",
                "muted",
            ),
        ]
    )
    cost_cards = "".join(
        [
            _kpi_card_html(
                "Стоимость месячного обязательства",
                money_ru(summary["total_value"]),
                "total",
            ),
            _kpi_card_html(
                "Стоимость допущенных работ",
                money_ru(summary["clean_value"]),
                "ready",
            ),
            _kpi_card_html(
                "Стоимость работ с риском",
                money_ru(summary["risk_value"]),
                "risk",
            ),
            _kpi_card_html(
                "Стоимость отложенных / исключённых",
                money_ru(deferred_value) if deferred_value > 0 else "—",
                "muted",
            ),
            _kpi_card_html("Доля риска", pct_display(summary["risk_share_pct"]), "risk"),
            _kpi_card_html(
                "Доля допущенного объёма",
                pct_display(summary["admitted_share_pct"]),
                "ready",
            ),
            _kpi_card_html(
                "Средняя стоимость BOQ-кода",
                money_ru(summary["avg_boq_value"]),
                "open",
            ),
        ]
    )
    labor_cards = "".join(
        [
            _kpi_card_html(
                "Требуемые чел·ч",
                kpi_hours_display(summary["total_hours"]),
                "total",
            ),
            _kpi_card_html(
                "Стоимость труда",
                money_ru(summary["total_labor"]),
                "total",
            ),
            _kpi_card_html(
                "Требуемые люди",
                kpi_people_display(summary["required_people"]),
                "open",
            ),
            _kpi_card_html(
                "Количество звеньев",
                str(summary["crew_count"]) if summary["crew_count"] else "—",
                "open",
            ),
            _kpi_card_html(
                "Средняя длительность, дн.",
                kpi_days_display(summary["avg_duration_days"]),
                "open",
            ),
            _kpi_card_html(
                "Средняя норма, чел·ч / ед.",
                kpi_norm_display(summary["avg_norm_hours"]),
                "open",
            ),
            _kpi_card_html(
                "Средняя выработка на смену",
                kpi_shift_output_display(summary["avg_shift_output"]),
                "open",
            ),
            _kpi_card_html(
                "Труд / стоимость работ, %",
                pct_display(summary["labor_pct"]),
                "risk",
            ),
        ]
    )

    st.markdown(
        f"""
        <div class="passport-kpi-panel">
            <div class="passport-kpi-group">
                <div class="passport-kpi-group-title">Объём месячного обязательства</div>
                <div class="v2-kpi-row v2-kpi-row--7">{volume_cards}</div>
            </div>
            <div class="passport-kpi-group">
                <div class="passport-kpi-group-title">Стоимость месячного обязательства</div>
                <div class="v2-kpi-row v2-kpi-row--7">{cost_cards}</div>
            </div>
            <div class="passport-kpi-group">
                <div class="passport-kpi-group-title">Трудозатраты месячного обязательства</div>
                <div class="v2-kpi-row v2-kpi-row--7">{labor_cards}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_passport_header(df: pd.DataFrame) -> None:
    if df.empty:
        return
    passport_status = (
        safe_str(df["passport_status"].iloc[0])
        if "passport_status" in df.columns
        else "—"
    )
    passport_name = (
        safe_str(df["passport_name"].iloc[0]) if "passport_name" in df.columns else "—"
    )
    approved_by = safe_str(df["approved_by"].iloc[0]) if "approved_by" in df.columns else "—"
    approved_at = safe_str(df["approved_at"].iloc[0]) if "approved_at" in df.columns else "—"
    passport_id = safe_str(df["passport_id"].iloc[0]) if "passport_id" in df.columns else "—"
    status_label = PASSPORT_STATUS_RU.get(passport_status, passport_status or "—")
    date_label = approved_at[:10] if approved_at and approved_at != "nan" else "—"
    id_label = passport_id if passport_id else "—"
    st.markdown(
        f"""
        <div class="passport-status-compact">
            <div class="passport-status-compact-title">Статус утверждённого паспорта</div>
            <div class="passport-status-grid">
                <div>
                    <div class="passport-status-item-label">Статус</div>
                    <div class="passport-status-item-value">{status_label}</div>
                </div>
                <div>
                    <div class="passport-status-item-label">Паспорт</div>
                    <div class="passport-status-item-value">{passport_name or "—"}</div>
                </div>
                <div>
                    <div class="passport-status-item-label">Утвердил</div>
                    <div class="passport-status-item-value">{approved_by or "—"}</div>
                </div>
                <div>
                    <div class="passport-status-item-label">Дата</div>
                    <div class="passport-status-item-value">{date_label}</div>
                </div>
                <div>
                    <div class="passport-status-item-label">passport_id</div>
                    <div class="passport-status-item-value">{id_label}</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_passport_filters(
    passport_df: pd.DataFrame,
    scope_ref: pd.DataFrame,
    v2_df: pd.DataFrame,
) -> Dict[str, str]:
    month_options = planning_month_filter_options(v2_df)
    project_options = planning_project_filter_options(scope_ref)
    title_options = planning_title_filter_options(scope_ref)
    discipline_options = planning_discipline_filter_options(scope_ref)
    system_options = planning_system_filter_options(scope_ref, v2_df)
    crew_options = planning_crew_filter_options(v2_df, passport_df)

    sync_selectbox_option(FILTER_KEYS["month"], month_options)
    sync_selectbox_option(FILTER_KEYS["project"], project_options)
    sync_selectbox_option(FILTER_KEYS["facility"], title_options)
    sync_selectbox_option(FILTER_KEYS["discipline"], discipline_options)
    sync_selectbox_option(FILTER_KEYS["system"], system_options)
    sync_selectbox_option(FILTER_KEYS["crew"], crew_options)
    sync_selectbox_option(FILTER_KEYS["queue"], QUEUE_FILTER_OPTIONS)

    with st.container(border=True):
        st.markdown('<div class="passport-filter-heading">Срез паспорта месяца</div>', unsafe_allow_html=True)
        r1c1, r1c2, r1c3, r1c4 = st.columns(4)
        r2c1, r2c2, r2c3, r2c4 = st.columns(4)
        r3c1, _, _, _ = st.columns(4)

        project = r1c1.selectbox(
            "Проект",
            project_options,
            key=FILTER_KEYS["project"],
        )
        month = r1c2.selectbox(
            "Месяц",
            month_options,
            key=FILTER_KEYS["month"],
        )
        queue = r1c3.selectbox(
            "Очередь",
            QUEUE_FILTER_OPTIONS,
            key=FILTER_KEYS["queue"],
        )
        facility = r1c4.selectbox(
            "Титул",
            title_options,
            key=FILTER_KEYS["facility"],
        )
        discipline = r2c1.selectbox(
            "Дисциплина",
            discipline_options,
            key=FILTER_KEYS["discipline"],
        )
        system = r2c2.selectbox(
            "Система",
            system_options,
            key=FILTER_KEYS["system"],
        )
        crew = r2c3.selectbox(
            "Звено",
            crew_options,
            key=FILTER_KEYS["crew"],
        )
        admission = r2c4.selectbox(
            "Статус допуска",
            filter_options(passport_df, "admission_status_label"),
            key=FILTER_KEYS["admission"],
        )
        risk = r3c1.selectbox(
            "Риск",
            RISK_FILTER_OPTIONS,
            key=FILTER_KEYS["risk"],
        )

    return {
        "project": project,
        "month": month,
        "queue": queue,
        "facility": facility,
        "discipline": discipline,
        "system": system,
        "crew": crew,
        "admission": admission,
        "risk": risk,
    }


def _xlsx_write_sheet(
    ws,
    title: str,
    headers: List[str],
    rows: List[List[Any]],
    *,
    header_fill: str = "1F3864",
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    title_fill = PatternFill("solid", fgColor="1F3864")
    header_fill_obj = PatternFill("solid", fgColor=header_fill)
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=10, color="404040")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    span = max(len(headers), 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center
    ws.row_dimensions[1].height = 24

    for col_idx, header in enumerate(headers, start=1):
        hcell = ws.cell(row=2, column=col_idx, value=header)
        hcell.font = header_font
        hcell.fill = header_fill_obj
        hcell.alignment = center

    for row_idx, row in enumerate(rows, start=3):
        for col_idx, value in enumerate(row, start=1):
            bcell = ws.cell(row=row_idx, column=col_idx, value=value)
            bcell.font = body_font
            bcell.alignment = left if col_idx <= 2 else center

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max(len(str(headers[col_idx - 1])) + 4, 12), 36)


def build_passport_xlsx_bytes(
    summary: Dict[str, Any],
    display_table: pd.DataFrame,
    breakdowns: Dict[str, pd.DataFrame],
    risk_table: pd.DataFrame,
    filters: Dict[str, str],
) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = _passport_summary_rows(summary, filters)
    ws_summary = wb.create_sheet("Сводка")
    _xlsx_write_sheet(ws_summary, "Паспорт месяца — сводка", ["Показатель", "Значение"], summary_rows)

    line_headers = list(display_table.columns) if not display_table.empty else ["Нет данных"]
    line_rows = display_table.values.tolist() if not display_table.empty else []
    ws_lines = wb.create_sheet("Строки паспорта")
    _xlsx_write_sheet(ws_lines, "Строки паспорта месяца", line_headers, line_rows)

    sheet_map = {
        "По дисциплинам": breakdowns.get("discipline", pd.DataFrame()),
        "По титулам": breakdowns.get("title", pd.DataFrame()),
        "По системам": breakdowns.get("system", pd.DataFrame()),
        "По звеньям": breakdowns.get("crew", pd.DataFrame()),
    }
    for sheet_name, bdf in sheet_map.items():
        ws = wb.create_sheet(sheet_name[:31])
        if bdf.empty:
            _xlsx_write_sheet(ws, sheet_name, ["Нет данных"], [])
        else:
            _xlsx_write_sheet(
                ws,
                sheet_name,
                list(bdf.columns),
                bdf.values.tolist(),
            )

    ws_risk = wb.create_sheet("Коды с риском")
    if risk_table.empty:
        _xlsx_write_sheet(ws_risk, "Коды с риском", ["Нет данных"], [])
    else:
        _xlsx_write_sheet(
            ws_risk,
            "Коды с риском",
            list(risk_table.columns),
            risk_table.values.tolist(),
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_export_filename(project: str, month: str) -> str:
    proj = project if project != "Все" else "ALL"
    mon = month if month != "Все" else "ALL"
    ts = datetime.now().strftime("%Y%m%d")
    return f"Monthly_Passport_{proj}_{mon}_{ts}.xlsx"


BREAKDOWN_GROUP_OPTIONS: Dict[str, tuple[str, str, str]] = {
    "Дисциплина": ("discipline_display", "Дисциплина", "discipline"),
    "Титул": ("title_display", "Титул", "facility"),
    "Система": ("system_label", "Система", "system"),
    "Звено": ("crew_label", "Звено", "crew"),
}


def _passport_summary_rows(summary: Dict[str, Any], filters: Dict[str, str]) -> List[List[Any]]:
    return [
        ["Строк в паспорте", summary.get("row_count", 0)],
        ["BOQ-кодов", summary.get("boq_count", 0)],
        ["Плановый объём", summary.get("planned_qty", 0)],
        ["Допущено", summary.get("clean_count", 0)],
        ["С риском", summary.get("risk_count", 0)],
        ["Стоимость обязательства", summary.get("total_value", 0)],
        ["Стоимость с риском", summary.get("risk_value", 0)],
        ["Трудозатраты, чел·ч", summary.get("total_hours", 0)],
        ["Стоимость труда", summary.get("total_labor", 0)],
        ["Труд / стоимость, %", summary.get("labor_pct", 0)],
        ["Проект", filters.get("project", "Все")],
        ["Месяц", filters.get("month", "Все")],
        ["Очередь", filters.get("queue", "Все")],
        ["Титул", filters.get("facility", "Все")],
        ["Дисциплина", filters.get("discipline", "Все")],
        [
            "Дата выгрузки",
            datetime.now(timezone.utc).astimezone().strftime("%d.%m.%Y %H:%M"),
        ],
    ]


def build_all_breakdowns(filtered: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    return {
        "discipline": build_breakdown_table(filtered, "discipline_display", "Дисциплина"),
        "title": build_breakdown_table(filtered, "title_display", "Титул"),
        "system": build_breakdown_table(filtered, "system_label", "Система"),
        "crew": build_breakdown_table(filtered, "crew_label", "Звено"),
    }


def render_slice_summary_block(filtered: pd.DataFrame, filters: Dict[str, str]) -> None:
    st.markdown("### Сводка по срезу")
    grouping = st.selectbox(
        "Группировка",
        list(BREAKDOWN_GROUP_OPTIONS),
        key="passport_slice_grouping",
    )
    group_col, label_col, filter_key = BREAKDOWN_GROUP_OPTIONS[grouping]
    breakdown_table = build_breakdown_table(filtered, group_col, label_col)
    if filters.get(filter_key, "Все") != "Все" and len(breakdown_table) <= 1:
        st.caption("Для сравнения групп сбросьте соответствующий фильтр.")
    if breakdown_table.empty:
        st.info("Нет данных для сводки по выбранной группировке.")
    else:
        st.dataframe(breakdown_table, use_container_width=True, hide_index=True)


def render_excel_exports(
    summary: Dict[str, Any],
    display_table: pd.DataFrame,
    breakdowns: Dict[str, pd.DataFrame],
    risk_table: pd.DataFrame,
    filters: Dict[str, str],
) -> None:
    st.markdown("### Excel-выгрузка")
    if not OPENPYXL_AVAILABLE:
        st.error("Для выгрузки Excel установите пакет: pip install openpyxl")
        return
    try:
        xlsx_bytes = build_passport_xlsx_bytes(
            summary,
            display_table,
            breakdowns,
            risk_table,
            filters,
        )
    except Exception as exc:  # noqa: BLE001
        st.error(f"Не удалось сформировать Excel: {exc}")
        return

    st.download_button(
        "Скачать Excel",
        data=xlsx_bytes,
        file_name=build_export_filename(filters["project"], filters["month"]),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=False,
    )


def render_documents_block() -> None:
    with st.expander("Документы: месяц / неделя / смена", expanded=False):
        st.caption(
            "Недельные и сменные задания будут использоваться как печатные/цифровые документы "
            "для мастеров. После выполнения факт будет сопоставляться с заданием: плановые объёмы, "
            "трудозатраты и стоимость труда → фактический прогресс → расчёт производительности → "
            "основа для умной заработной платы."
        )

        st.markdown("**Паспорт месяца**")
        st.caption(
            "Фиксация утверждённого месячного обязательства: объёмы, стоимость, трудозатраты, "
            "звенья, риски."
        )
        m1, _ = st.columns([1, 3])
        with m1:
            st.button(
                "PDF: Паспорт месяца",
                disabled=True,
                help="Следующий этап",
                key="passport_doc_pdf_month",
            )

        st.markdown("**Недельное задание**")
        st.caption(
            "Документ для передачи мастерам: что делаем на неделе, где делаем, какими звеньями, "
            "с какими объёмами, трудозатратами, материалами, схемами и ограничениями."
        )
        w1, w2, _ = st.columns([1, 1, 2])
        with w1:
            st.selectbox("Неделя", ["—"], disabled=True, key="passport_doc_week")
        with w2:
            st.button(
                "PDF: Недельное задание",
                disabled=True,
                help="Следующий этап",
                key="passport_doc_pdf_week",
            )

        st.markdown("**Сменное наряд-задание**")
        st.caption(
            "Операционное задание на смену: локация, фрагменты РД/схемы, BOQ-коды, IWP/пакеты, "
            "системы, спецификация МТР, объёмы, состав звена, плановые чел·ч, стоимость труда "
            "и ожидаемый результат."
        )
        s1, s2, s3, _ = st.columns([1, 1, 1, 1])
        with s1:
            st.date_input(
                "Дата смены",
                value=datetime.now().date(),
                disabled=True,
                key="passport_doc_shift_date",
            )
        with s2:
            st.selectbox("Звено", ["—"], disabled=True, key="passport_doc_shift_crew")
        with s3:
            st.button(
                "PDF: Сменное наряд-задание",
                disabled=True,
                help="Следующий этап",
                key="passport_doc_pdf_shift",
            )


def render_passport_page(
    passport_df: pd.DataFrame,
    scope_ref: pd.DataFrame,
    v2_df: pd.DataFrame,
) -> None:
    inject_passport_page_styles()
    filters = render_passport_filters(passport_df, scope_ref, v2_df)
    filtered = apply_passport_filters(passport_df, **filters)

    if filtered.empty:
        st.info(NO_PASSPORT_FOR_SLICE_TEXT)
        return

    render_passport_header(filtered)

    summary = compute_passport_summary(filtered)
    st.markdown("### Сводка месячного обязательства")
    render_passport_summary_panel(summary)

    display_table = build_passport_display_table(filtered)
    st.markdown("### Строки паспорта месяца")
    st.caption(f"Показано {len(display_table)} строк.")
    st.dataframe(display_table, use_container_width=True, hide_index=True, height=min(680, 80 + 35 * len(display_table)))

    render_slice_summary_block(filtered, filters)
    breakdowns = build_all_breakdowns(filtered)

    risk_df = filtered[filtered["is_risk"].astype(bool)] if "is_risk" in filtered.columns else pd.DataFrame()
    risk_table = build_passport_display_table(risk_df)

    render_excel_exports(summary, display_table, breakdowns, risk_table, filters)
    render_documents_block()


st.title("Паспорт месяца")
st.caption(
    "Итоговая витрина утверждённого месячного обязательства после War Room. "
    "Фильтры → сводка → таблица → разбивки → Excel."
)

passport_raw = load_passport_dataset()
v2_df = load_v2_plan_lines()
passport_df = enrich_passport_dataframe(passport_raw, v2_df)
scope_ref = load_planning_scope_filter_reference()

render_passport_page(passport_df, scope_ref, v2_df)
