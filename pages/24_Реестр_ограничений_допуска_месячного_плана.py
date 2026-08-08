"""
Реестр ограничений допуска месячного плана — UI (R1 + R2 update form).

Источник: services.monthly_plan_constraint_registry_service
R2: форма «Обновить ограничение» (OPEN/IN_PROGRESS) + карточка/фильтры/история UPDATED.
Снятие: RPC resolve_monthly_plan_constraint (отдельная форма).
"""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

import pandas as pd
import streamlit as st

from services.monthly_plan_constraint_registry_service import (
    CONSTRAINT_PRIORITY_DISPLAY,
    DEADLINE_SOURCE_DISPLAY,
    DEADLINE_STATUS_DISPLAY,
    UPDATE_PATCH_WHITELIST,
    build_registry_summary,
    clear_constraint_registry_caches,
    clear_registry_read_caches,
    display_constraint_priority,
    display_deadline_source,
    display_deadline_status,
    load_constraint_events,
    load_constraint_registry,
    normalize_update_patch,
    parse_constraint_event_payload,
    resolve_constraint,
    update_constraint,
)

st.set_page_config(layout="wide", page_title="Реестр ограничений производства работ")

ALL = "Все"
OPEN_ONLY = "Только открытые"
MODE_ALL = "Все"
OPEN_RESOLUTIONS = frozenset({"OPEN", "IN_PROGRESS"})
CLOSED_RESOLUTIONS = frozenset({"RESOLVED", "CANCELLED"})

# Hide fixture projects from user-facing filter options (data unchanged).
HIDDEN_PROJECT_CODES = frozenset({"TEST_REG", "TEST_REG_R2"})

EXPORT_SCOPE_FILTERED = "По текущим фильтрам"
EXPORT_SCOPE_FULL_PROJECT = "Весь проект — все месяцы и дисциплины"
EXPORT_SCOPE_OPTIONS = [EXPORT_SCOPE_FILTERED, EXPORT_SCOPE_FULL_PROJECT]

# Excel export columns: (source field, Russian header, cell kind).
# kind: text | date | int | money | check | resolution | priority | deadline | source | author | recorder
EXPORT_COLUMN_SPECS: tuple[tuple[str, str, str], ...] = (
    ("project_code", "Проект", "text"),
    ("month_key", "Месяц", "text"),
    ("queue", "Очередь", "text"),
    ("facility_building", "Титул", "text"),
    ("construction_discipline", "Дисциплина", "text"),
    ("boq_code", "BOQ", "text"),
    ("boq_name", "Наименование работ", "text"),
    ("work_package", "Пакет работ", "text"),
    ("system", "Система", "text"),
    ("unit", "Ед. изм.", "text"),
    ("planned_qty", "Плановый объём", "qty"),
    ("plan_value", "Плановая стоимость", "money"),
    ("responsible_department", "Отдел", "text"),
    ("check_status", "Статус проверки", "check"),
    ("resolution_status", "Статус устранения", "resolution"),
    ("constraint_category", "Тип ограничения", "text"),
    ("constraint_priority", "Приоритет", "priority"),
    ("constraint_occurred_at", "Дата возникновения", "date"),
    ("constraint_created_at", "Дата регистрации", "date"),
    ("days_open_real", "Дней открыто", "int"),
    ("problem_summary", "Суть ограничения", "text"),
    ("problem_impact", "Влияние", "author"),
    ("required_action", "Требуемое действие", "author"),
    ("problem_owner", "Владелец ограничения", "author"),
    ("owner_name", "Исполнитель устранения", "author"),
    ("subcontractor_coordinator", "Координатор от субподрядчика", "author"),
    ("created_by", "Зафиксировал ограничение", "recorder"),
    ("deadline_status", "Статус срока", "deadline"),
    ("deadline_source", "Источник срока", "source"),
    ("target_resolution_date", "Плановая дата устранения", "date"),
    ("next_control_date", "Следующий контроль", "date"),
    ("actual_resolution_date", "Фактическая дата устранения", "date"),
    ("value_at_risk", "Стоимость под риском", "money"),
    ("comment", "Примечание", "text"),
    ("updated_by", "Последнее изменение", "text"),
    ("resolved_by", "Кто снял", "text"),
    ("last_action_at", "Дата последнего действия", "date"),
    ("constraint_id", "ID ограничения", "text"),
)

EXPORT_WRAP_HEADERS = frozenset(
    {
        "Наименование работ",
        "Суть ограничения",
        "Влияние",
        "Требуемое действие",
        "Примечание",
        "Пакет работ",
        "Система",
    }
)
EXPORT_SHEET_REGISTRY = "Реестр ограничений"
EXPORT_SHEET_PARAMS = "Параметры выгрузки"

# Enterprise XLSX visual style (presentation only; openpyxl Color RGB).
COLOR_NAVY = "17365D"
COLOR_BLUE = "2F5F9F"
COLOR_LIGHT_BLUE = "D6E3F0"
COLOR_META_BLUE = "E8F0F8"
COLOR_KPI_BG = "F2F5F9"
COLOR_ALT_ROW = "F7FAFC"
COLOR_BORDER = "B8C7D9"
COLOR_WHITE = "FFFFFF"
COLOR_HOLD = "FCE4D6"
COLOR_FAIL = "F8CBAD"
COLOR_WARNING = "FFF2CC"
COLOR_OPEN = "DDEBF7"
COLOR_IN_PROGRESS = "E2EFDA"
COLOR_RESOLVED = "C6EFCE"
COLOR_CANCELLED = "D9D9D9"
COLOR_CRITICAL = "F8CBAD"
COLOR_HIGH = "FCE4D6"
COLOR_NORMAL = "DDEBF7"
COLOR_OVERDUE = "F8CBAD"
FONT_NAME = "Calibri"

EXPORT_TITLE = "РЕЕСТР ОГРАНИЧЕНИЙ ПРОИЗВОДСТВА РАБОТ"
EXPORT_SUBTITLE = (
    "Единый реестр ограничений, препятствующих допуску, "
    "началу или продолжению работ месячного плана."
)

# Layout rows on registry sheet (1-based).
EXPORT_ROW_TITLE = 1
EXPORT_ROW_SUBTITLE = 2
EXPORT_ROW_META = 3
EXPORT_ROW_KPI_LABEL = 4
EXPORT_ROW_KPI_VALUE = 5
EXPORT_ROW_TABLE_HEADER = 7
EXPORT_ROW_DATA_START = 8

EXPORT_COLUMN_WIDTHS: dict[str, float] = {
    "Проект": 16,
    "Месяц": 14,
    "Очередь": 12,
    "Титул": 15,
    "Дисциплина": 22,
    "BOQ": 18,
    "Наименование работ": 36,
    "Пакет работ": 20,
    "Система": 18,
    "Ед. изм.": 10,
    "Плановый объём": 13,
    "Плановая стоимость": 16,
    "Отдел": 18,
    "Статус проверки": 18,
    "Статус устранения": 16,
    "Тип ограничения": 22,
    "Приоритет": 14,
    "Дата возникновения": 14,
    "Дата регистрации": 14,
    "Дней открыто": 12,
    "Суть ограничения": 40,
    "Влияние": 36,
    "Требуемое действие": 36,
    "Владелец ограничения": 22,
    "Исполнитель устранения": 22,
    "Координатор от субподрядчика": 24,
    "Зафиксировал ограничение": 22,
    "Статус срока": 16,
    "Источник срока": 16,
    "Плановая дата устранения": 16,
    "Следующий контроль": 14,
    "Фактическая дата устранения": 16,
    "Стоимость под риском": 16,
    "Примечание": 32,
    "Последнее изменение": 20,
    "Кто снял": 16,
    "Дата последнего действия": 16,
    "ID ограничения": 22,
}

# Pastel fills for status / priority / deadline cells only (display text → RGB).
EXPORT_STATUS_FILLS: dict[str, str] = {
    "Удержание": COLOR_HOLD,
    "HOLD": COLOR_HOLD,
    "Заблокировано": COLOR_FAIL,
    "FAIL": COLOR_FAIL,
    "Требует уточнения": COLOR_WARNING,
    "WARNING": COLOR_WARNING,
    "Ожидает проверки": COLOR_OPEN,
    "ОЖИДАЕТ": COLOR_OPEN,
    "Проверка пройдена": COLOR_RESOLVED,
    "PASS": COLOR_RESOLVED,
    "Открыто": COLOR_OPEN,
    "OPEN": COLOR_OPEN,
    "В работе": COLOR_IN_PROGRESS,
    "IN_PROGRESS": COLOR_IN_PROGRESS,
    "Снято": COLOR_RESOLVED,
    "RESOLVED": COLOR_RESOLVED,
    "Отменено": COLOR_CANCELLED,
    "CANCELLED": COLOR_CANCELLED,
    "Критический": COLOR_CRITICAL,
    "CRITICAL": COLOR_CRITICAL,
    "Высокий": COLOR_HIGH,
    "HIGH": COLOR_HIGH,
    "Нормальный": COLOR_NORMAL,
    "NORMAL": COLOR_NORMAL,
    "Низкий": COLOR_NORMAL,
    "LOW": COLOR_NORMAL,
    "Просрочен": COLOR_OVERDUE,
    "Просрочено": COLOR_OVERDUE,
    "OVERDUE": COLOR_OVERDUE,
    "Исполнен": COLOR_RESOLVED,
    "DONE": COLOR_RESOLVED,
}

# UI-only calendar months for Page 24 filter (no test months).
MONTH_OPTIONS_2026 = [
    "январь-2026",
    "февраль-2026",
    "март-2026",
    "апрель-2026",
    "май-2026",
    "июнь-2026",
    "июль-2026",
    "август-2026",
    "сентябрь-2026",
    "октябрь-2026",
    "ноябрь-2026",
    "декабрь-2026",
]

# Product-canonical construction_discipline (already used in monthly_plan_constraints).
PIPE_TESTING_DISCIPLINE = "Испытания трубопроводов"

ADMISSION_OUTCOME_MESSAGES = {
    "READY": "BOQ допущен и доступен для включения в месячный паспорт",
    "READY_WITH_RISK": "Ограничение снято, но по BOQ остаются риски",
    "BLOCKED": "Ограничение снято, но по BOQ остаются блокирующие ограничения",
    "WAITING": "Ограничение снято, но BOQ ожидает проверки/решения других отделов",
}

CHECK_STATUS_OPTIONS = ["ОЖИДАЕТ", "PASS", "WARNING", "HOLD", "FAIL"]
RESOLUTION_STATUS_OPTIONS = ["OPEN", "IN_PROGRESS", "RESOLVED", "CANCELLED"]

# Display-only labels (filters / card; DB keep technical values)
CHECK_STATUS_DISPLAY = {
    "ОЖИДАЕТ": "Ожидает проверки",
    "WARNING": "Требует уточнения",
    "HOLD": "Удержание",
    "FAIL": "Заблокировано",
    "PASS": "Проверка пройдена",
}
RESOLUTION_STATUS_DISPLAY = {
    "OPEN": "Открыто",
    "IN_PROGRESS": "В работе",
    "RESOLVED": "Снято",
    "CANCELLED": "Отменено",
}

# Table labels: uppercase (visual registry table)
CHECK_STATUS_TABLE_DISPLAY = {
    "ОЖИДАЕТ": "ОЖИДАЕТ ПРОВЕРКИ",
    "WARNING": "ТРЕБУЕТ УТОЧНЕНИЯ",
    "HOLD": "УДЕРЖАНИЕ",
    "FAIL": "ЗАБЛОКИРОВАНО",
    "PASS": "ПРОВЕРКА ПРОЙДЕНА",
}
RESOLUTION_STATUS_TABLE_DISPLAY = {
    "OPEN": "ОТКРЫТО",
    "IN_PROGRESS": "В РАБОТЕ",
    "RESOLVED": "СНЯТО",
    "CANCELLED": "ОТМЕНЕНО",
}
PRIORITY_TABLE_DISPLAY = {
    "CRITICAL": "КРИТИЧЕСКИЙ",
    "HIGH": "ВЫСОКИЙ",
    "NORMAL": "НОРМАЛЬНЫЙ",
    "LOW": "НИЗКИЙ",
}

# Text-only styles for pandas Styler on st.dataframe (color + bold; no background).
# Same renderer family as Page 23 «Единый реестр кодов месяца».
CHECK_STATUS_TEXT_STYLE = {
    "ОЖИДАЕТ ПРОВЕРКИ": "color:#6B7280;font-weight:700;",
    "ТРЕБУЕТ УТОЧНЕНИЯ": "color:#D97706;font-weight:700;",
    "УДЕРЖАНИЕ": "color:#EA580C;font-weight:700;",
    "ЗАБЛОКИРОВАНО": "color:#DC2626;font-weight:700;",
    "ПРОВЕРКА ПРОЙДЕНА": "color:#15803D;font-weight:700;",
}
RESOLUTION_STATUS_TEXT_STYLE = {
    "ОТКРЫТО": "color:#475569;font-weight:700;",
    "В РАБОТЕ": "color:#2563EB;font-weight:700;",
    "СНЯТО": "color:#15803D;font-weight:700;",
    "ОТМЕНЕНО": "color:#6B7280;font-weight:700;",
}
PRIORITY_TEXT_STYLE = {
    "КРИТИЧЕСКИЙ": "color:#DC2626;font-weight:700;",
    "ВЫСОКИЙ": "color:#EA580C;font-weight:700;",
    "НОРМАЛЬНЫЙ": "color:#2563EB;font-weight:700;",
    "НИЗКИЙ": "color:#6B7280;font-weight:700;",
}

EMPTY_AUTHOR = "Не заполнено автором"
EMPTY_RECORDER = "Не зафиксировано"
COMMENT_TABLE_MAX_LEN = 80

DATE_MODE_KEEP = "Не менять"
DATE_MODE_SET = "Указать дату"
DATE_MODE_CLEAR = "Очистить"
DATE_MODE_OPTIONS = [DATE_MODE_KEEP, DATE_MODE_SET, DATE_MODE_CLEAR]

PRIORITY_FILTER_OPTIONS = list(CONSTRAINT_PRIORITY_DISPLAY.keys())
DEADLINE_STATUS_FILTER_OPTIONS = list(DEADLINE_STATUS_DISPLAY.keys())

UPDATE_TEXT_FIELDS = (
    "problem_owner",
    "owner_name",
    "subcontractor_coordinator",
    "constraint_category",
    "problem_description",
    "problem_impact",
    "required_action",
)

UPDATE_ENUM_FIELDS = (
    "constraint_priority",
    "deadline_status",
    "deadline_source",
)

UPDATE_DATE_FIELDS = (
    "constraint_occurred_at",
    "target_resolution_date",
    "next_control_date",
)

FIELD_LABELS_RU = {
    "constraint_occurred_at": "Дата возникновения ограничения",
    "problem_owner": "Владелец ограничения",
    "owner_name": "Исполнитель устранения",
    "subcontractor_coordinator": "Координатор от субподрядчика",
    "constraint_category": "Категория ограничения",
    "constraint_priority": "Приоритет",
    "problem_description": "Подробное описание проблемы",
    "problem_impact": "Влияние",
    "required_action": "Требуемое действие",
    "deadline_status": "Статус срока",
    "deadline_source": "Источник срока",
    "target_resolution_date": "Плановая дата устранения",
    "next_control_date": "Дата следующего контроля",
}

TABLE_COLUMNS = [
    ("project_code", "Проект", 100),
    ("month_key", "Месяц", 110),
    ("constraint_occurred_at", "Дата возникновения", 110),
    ("constraint_created_at", "Дата фиксации", 110),
    ("queue", "Очередь", 90),
    ("facility_building", "Титул", 110),
    ("construction_discipline", "Дисциплина", 140),
    ("boq_code", "BOQ-код", 130),
    ("boq_name", "Наименование работ", 200),
    ("work_package", "Пакет работ / IWP", 130),
    ("system", "Система", 130),
    ("unit", "Ед. изм.", 70),
    ("planned_qty", "Плановый объём", 100),
    ("plan_value", "Плановая стоимость", 120),
    ("responsible_department", "Отдел", 110),
    ("check_status", "Статус проверки", 160),
    ("resolution_status", "Статус устранения", 130),
    ("constraint_category", "Тип ограничения", 120),
    ("constraint_priority", "Приоритет", 120),
    ("problem_summary", "Суть ограничения", 200),
    ("problem_owner", "Владелец ограничения", 150),
    ("owner_name", "Исполнитель", 150),
    ("subcontractor_coordinator", "Координатор", 150),
    ("problem_impact", "Влияние", 200),
    ("required_action", "Требуемое действие", 200),
    ("deadline_status", "Статус срока", 120),
    ("target_resolution_date", "Плановая дата устранения", 120),
    ("next_control_date", "Следующий контроль", 120),
    ("actual_resolution_date", "Фактическая дата устранения", 120),
    ("days_open_real", "Дней открыто", 90),
    ("delay_days", "Задержка, дней", 90),
    ("value_at_risk", "Стоимость под риском", 120),
    ("comment", "Примечание", 200),
    ("created_by", "Зафиксировал ограничение", 150),
    ("updated_by", "Последнее изменение", 130),
    ("resolved_by", "Кто снял", 120),
    ("last_action_at", "Дата последнего действия", 130),
]

# Same height family as Page 23 «Единый реестр кодов месяца».
REGISTRY_TABLE_HEIGHT_PX = 36 * 25 + 38
REGISTRY_SELECT_KEY = "reg_dataframe"
REGISTRY_SESSION_SELECTED = "reg_selected_constraint_id"


def safe_str(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "<na>"}:
        return ""
    return text


def safe_num(value: Any) -> float:
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def safe_date(value: Any) -> Optional[date]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = safe_str(value)
    if not text:
        return None
    try:
        parsed = pd.to_datetime(text, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    except Exception:  # noqa: BLE001
        return None


def format_date_ru(value: Any) -> str:
    d = safe_date(value)
    return d.strftime("%d.%m.%Y") if d else "—"


def format_datetime_ru(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "—"
    try:
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return display_dash(value)
        if getattr(parsed, "tzinfo", None) is not None:
            parsed = parsed.tz_convert(None)
        return parsed.strftime("%d.%m.%Y %H:%M")
    except Exception:  # noqa: BLE001
        return display_dash(value)


def money_ru(value: Any) -> str:
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return "—"
        amount = float(value)
        sign = "-" if amount < 0 else ""
        amount = abs(amount)
        whole, frac = f"{amount:.2f}".split(".")
        whole_fmt = f"{int(whole):,}".replace(",", " ")
        return f"{sign}{whole_fmt},{frac} ₽"
    except Exception:  # noqa: BLE001
        return "—"


def qty_ru(value: Any) -> str:
    try:
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return "—"
        return f"{float(value):,.2f}".replace(",", " ").replace(".", ",")
    except Exception:  # noqa: BLE001
        return "—"


def display_dash(value: Any) -> str:
    text = safe_str(value)
    return text if text else "—"


def display_author_empty(value: Any) -> str:
    """UI-only empty marker; never written to DB."""
    text = safe_str(value)
    return text if text else EMPTY_AUTHOR


def display_recorder(value: Any) -> str:
    """created_by display: who first recorded the constraint."""
    text = safe_str(value)
    return text if text else EMPTY_RECORDER


def display_updated_by(value: Any) -> str:
    """updated_by display: who last changed the record."""
    text = safe_str(value)
    return text if text else "—"


def display_resolved_by(value: Any) -> str:
    """resolved_by display: who closed the constraint."""
    text = safe_str(value)
    return text if text else "—"


def format_date_author_empty(value: Any) -> str:
    d = safe_date(value)
    return d.strftime("%d.%m.%Y") if d else EMPTY_AUTHOR


def display_check_status(value: Any) -> str:
    key = safe_str(value).upper()
    if not key:
        return "—"
    return CHECK_STATUS_DISPLAY.get(key, safe_str(value))


def display_resolution_status(value: Any) -> str:
    key = safe_str(value).upper()
    if not key:
        return "—"
    return RESOLUTION_STATUS_DISPLAY.get(key, safe_str(value))


def display_check_status_table(value: Any) -> str:
    key = safe_str(value).upper()
    if not key:
        return "—"
    return CHECK_STATUS_TABLE_DISPLAY.get(key, safe_str(value).upper())


def display_resolution_status_table(value: Any) -> str:
    key = safe_str(value).upper()
    if not key:
        return "—"
    return RESOLUTION_STATUS_TABLE_DISPLAY.get(key, safe_str(value).upper())


def display_priority_table(value: Any) -> str:
    key = safe_str(value).upper()
    if not key:
        return "—"
    return PRIORITY_TABLE_DISPLAY.get(key, safe_str(value).upper())


def format_check_filter_option(opt: str) -> str:
    if opt == ALL:
        return ALL
    return CHECK_STATUS_DISPLAY.get(opt, opt)


def format_resolution_filter_option(opt: str) -> str:
    if opt == ALL:
        return ALL
    return RESOLUTION_STATUS_DISPLAY.get(opt, opt)


def truncate_comment(value: Any, max_len: int = COMMENT_TABLE_MAX_LEN) -> str:
    text = safe_str(value)
    if not text:
        return "—"
    if len(text) <= max_len:
        return text
    return text[: max(0, max_len - 1)] + "…"


def _cell_text_style(styles: dict[str, str], value: Any) -> str:
    return styles.get(str(value).strip(), "")


def style_registry_display(df: pd.DataFrame):
    """
    Text-only styles for status/priority columns (no background).
    Compatible with st.dataframe + pandas Styler (same family as Page 23).
    """
    styler = df.style

    def style_check(val: Any) -> str:
        return _cell_text_style(CHECK_STATUS_TEXT_STYLE, val)

    def style_resolution(val: Any) -> str:
        return _cell_text_style(RESOLUTION_STATUS_TEXT_STYLE, val)

    def style_priority(val: Any) -> str:
        return _cell_text_style(PRIORITY_TEXT_STYLE, val)

    mapping = (
        ("Статус проверки", style_check),
        ("Статус устранения", style_resolution),
        ("Приоритет", style_priority),
    )
    for col, fn in mapping:
        if col not in df.columns:
            continue
        if hasattr(styler, "map"):
            styler = styler.map(fn, subset=pd.IndexSlice[:, [col]])
        else:
            styler = styler.applymap(fn, subset=pd.IndexSlice[:, [col]])
    return styler


def resolve_registry_selection(constraint_ids: list[str]) -> Optional[str]:
    """
    Page 23-style selection: dataframe rows → session → first available.
    """
    if not constraint_ids:
        return None

    selected_id: Optional[str] = None
    sel_state = st.session_state.get(REGISTRY_SELECT_KEY)
    if isinstance(sel_state, dict):
        rows_sel = sel_state.get("selection", {}).get("rows", [])
        if rows_sel:
            idx = int(rows_sel[0])
            if 0 <= idx < len(constraint_ids):
                selected_id = constraint_ids[idx]

    if not selected_id:
        stored = safe_str(st.session_state.get(REGISTRY_SESSION_SELECTED))
        if stored in constraint_ids:
            selected_id = stored

    if not selected_id:
        selected_id = constraint_ids[0]

    st.session_state[REGISTRY_SESSION_SELECTED] = selected_id
    return selected_id


def unique_sorted(series: pd.Series) -> list[str]:
    values = {safe_str(v) for v in series.dropna().tolist() if safe_str(v)}
    return sorted(values)


def risk_value(row: pd.Series) -> float:
    raw = row.get("value_at_risk")
    if raw is not None and not (isinstance(raw, float) and pd.isna(raw)):
        return safe_num(raw)
    return safe_num(row.get("plan_value"))


def can_edit_update_form(resolution_status: Any) -> bool:
    return safe_str(resolution_status).upper() in OPEN_RESOLUTIONS


def render_update_date_controls() -> tuple[dict[str, str], dict[str, Any]]:
    """
    Nullable date mode + date_input widgets outside st.form.

    Must stay outside the form so switching
    «Не менять / Указать дату / Очистить» triggers an immediate rerun
    and date_input appears without submit.
    """
    specs = (
        (
            "constraint_occurred_at",
            "Дата возникновения ограничения",
            "Указать дату возникновения",
        ),
        (
            "target_resolution_date",
            "Плановая дата устранения",
            "Указать плановую дату",
        ),
        (
            "next_control_date",
            "Дата следующего контроля",
            "Указать дату контроля",
        ),
    )
    modes: dict[str, str] = {}
    values: dict[str, Any] = {}
    cols = st.columns(3)
    for col, (field, mode_label, date_label) in zip(cols, specs):
        with col:
            mode = st.selectbox(
                mode_label,
                options=DATE_MODE_OPTIONS,
                key=f"reg_upd_{field}_mode",
            )
            modes[field] = mode
            values[field] = None
            if mode == DATE_MODE_SET:
                values[field] = st.date_input(
                    date_label,
                    key=f"reg_upd_{field}_date",
                )
    return modes, values


def _norm_comparable_text(value: Any) -> Optional[str]:
    text = safe_str(value)
    if not text or text in {"—", "-", "–"}:
        return None
    return text


def _norm_comparable_date(value: Any) -> Optional[str]:
    d = safe_date(value)
    return d.isoformat() if d else None


def row_to_update_baseline(row: pd.Series) -> dict[str, Any]:
    """Baseline values for dirty-patch comparison (technical codes / ISO dates)."""
    baseline: dict[str, Any] = {}
    for field in UPDATE_TEXT_FIELDS:
        baseline[field] = _norm_comparable_text(row.get(field))
    for field in UPDATE_ENUM_FIELDS:
        code = safe_str(row.get(field)).upper() or None
        baseline[field] = code
    for field in UPDATE_DATE_FIELDS:
        baseline[field] = _norm_comparable_date(row.get(field))
    return baseline


def build_update_dirty_patch(
    baseline: dict[str, Any],
    current: dict[str, Any],
) -> dict[str, Any]:
    """
    Dirty patch for update RPC:
    - unchanged → key omitted
    - changed → key=value
    - clear → key=null
    Does not auto-send null for untouched fields.
    """
    patch: dict[str, Any] = {}
    for field in UPDATE_TEXT_FIELDS + UPDATE_ENUM_FIELDS + UPDATE_DATE_FIELDS:
        if field not in current:
            continue
        if field not in UPDATE_PATCH_WHITELIST:
            continue
        new_val = current[field]
        old_val = baseline.get(field)
        if new_val == old_val:
            continue
        patch[field] = new_val
    return patch


def hydrate_update_form_state(
    session_state: Any,
    constraint_id: str,
    baseline: dict[str, Any],
) -> bool:
    """
    Namespace update form by constraint_id.
    Returns True if hydrated (constraint changed), False if same constraint (keep live input).
    """
    cid = safe_str(constraint_id)
    prev = safe_str(session_state.get("reg_upd_active_cid"))
    if prev == cid and session_state.get("reg_upd_baseline") is not None:
        return False

    session_state["reg_upd_active_cid"] = cid
    session_state["reg_upd_baseline"] = dict(baseline)

    for field in UPDATE_TEXT_FIELDS:
        session_state[f"reg_upd_{field}"] = baseline.get(field) or ""

    for field in UPDATE_ENUM_FIELDS:
        code = baseline.get(field) or ""
        session_state[f"reg_upd_{field}"] = code

    for field in UPDATE_DATE_FIELDS:
        session_state[f"reg_upd_{field}_mode"] = DATE_MODE_KEEP
        d = safe_date(baseline.get(field))
        session_state[f"reg_upd_{field}_date"] = d or date.today()

    # Keep actor/comment across rows only if empty; otherwise reset comment
    if "reg_upd_updated_by" not in session_state:
        session_state["reg_upd_updated_by"] = ""
    session_state["reg_upd_update_comment"] = ""
    return True


def collect_update_form_current(
    *,
    text_values: dict[str, Any],
    enum_values: dict[str, Any],
    date_modes: dict[str, str],
    date_values: dict[str, Any],
) -> dict[str, Any]:
    """Build current technical values from form inputs (including date modes)."""
    current: dict[str, Any] = {}
    for field, raw in text_values.items():
        current[field] = _norm_comparable_text(raw)
    for field, raw in enum_values.items():
        code = safe_str(raw).upper()
        current[field] = code or None
    for field in UPDATE_DATE_FIELDS:
        mode = date_modes.get(field, DATE_MODE_KEEP)
        if mode == DATE_MODE_KEEP:
            continue  # omit from current → dirty patch skips
        if mode == DATE_MODE_CLEAR:
            current[field] = None
        else:
            current[field] = _norm_comparable_date(date_values.get(field))
    return current


def format_update_value_display(field: str, value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return EMPTY_AUTHOR
    if field == "deadline_status":
        return display_deadline_status(value)
    if field == "deadline_source":
        return display_deadline_source(value)
    if field == "constraint_priority":
        return display_constraint_priority(value)
    if field in UPDATE_DATE_FIELDS:
        return format_date_author_empty(value)
    text = safe_str(value)
    return text if text else EMPTY_AUTHOR


def apply_filters(
    df: pd.DataFrame,
    *,
    project: str,
    month: str,
    facility: str,
    discipline: str,
    department: str,
    check_status: str,
    resolution_status: str,
    open_mode: str,
    search: str,
    problem_owner: str = ALL,
    owner_name: str = ALL,
    subcontractor_coordinator: str = ALL,
    constraint_priority: str = ALL,
    effective_deadline_status: str = ALL,
    is_deadline_overdue: str = ALL,
    is_next_control_overdue: str = ALL,
    queue: str = ALL,
) -> pd.DataFrame:
    if df.empty:
        return df

    result = df.copy()

    if project != ALL and "project_code" in result.columns:
        result = result[result["project_code"].astype(str) == project]
    if month != ALL and "month_key" in result.columns:
        result = result[result["month_key"].astype(str) == month]
    if facility != ALL and "facility_building" in result.columns:
        result = result[result["facility_building"].astype(str) == facility]
    if discipline != ALL and "construction_discipline" in result.columns:
        result = result[result["construction_discipline"].astype(str) == discipline]
    if department != ALL and "responsible_department" in result.columns:
        result = result[result["responsible_department"].astype(str) == department]
    if queue != ALL and "queue" in result.columns:
        result = result[result["queue"].astype(str).str.strip() == queue.strip()]
    if check_status != ALL and "check_status" in result.columns:
        result = result[
            result["check_status"].astype(str).str.strip().str.upper()
            == check_status.strip().upper()
        ]
    if resolution_status != ALL and "resolution_status" in result.columns:
        result = result[
            result["resolution_status"].astype(str).str.strip().str.upper()
            == resolution_status.strip().upper()
        ]

    # Состояние реестра: Все = no extra filter; Только открытые = OPEN+IN_PROGRESS
    if open_mode == OPEN_ONLY and "resolution_status" in result.columns:
        res = result["resolution_status"].astype(str).str.strip().str.upper()
        result = result[res.isin(OPEN_RESOLUTIONS)]

    if problem_owner != ALL and "problem_owner" in result.columns:
        result = result[
            result["problem_owner"].astype(str).str.strip() == problem_owner.strip()
        ]
    if owner_name != ALL and "owner_name" in result.columns:
        result = result[
            result["owner_name"].astype(str).str.strip() == owner_name.strip()
        ]
    if subcontractor_coordinator != ALL and "subcontractor_coordinator" in result.columns:
        result = result[
            result["subcontractor_coordinator"].astype(str).str.strip()
            == subcontractor_coordinator.strip()
        ]
    if constraint_priority != ALL and "constraint_priority" in result.columns:
        result = result[
            result["constraint_priority"].astype(str).str.strip().str.upper()
            == constraint_priority.strip().upper()
        ]
    if effective_deadline_status != ALL and "effective_deadline_status" in result.columns:
        result = result[
            result["effective_deadline_status"].astype(str).str.strip().str.upper()
            == effective_deadline_status.strip().upper()
        ]

    def _bool_filter(col: str, mode: str) -> None:
        nonlocal result
        if mode == ALL or col not in result.columns:
            return
        flags = result[col].fillna(False).astype(bool)
        if mode == "Да":
            result = result[flags]
        elif mode == "Нет":
            result = result[~flags]

    _bool_filter("is_deadline_overdue", is_deadline_overdue)
    _bool_filter("is_next_control_overdue", is_next_control_overdue)

    q = search.strip().lower()
    if q:
        boq = (
            result["boq_code"].astype(str).str.lower()
            if "boq_code" in result.columns
            else pd.Series("", index=result.index)
        )
        name = (
            result["boq_name"].astype(str).str.lower()
            if "boq_name" in result.columns
            else pd.Series("", index=result.index)
        )
        result = result[boq.str.contains(q, na=False) | name.str.contains(q, na=False)]

    return result


def sort_registry_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    result = df.copy()
    result["_sort_block"] = (
        (~result["is_blocking"].fillna(False).astype(bool)).astype(int)
        if "is_blocking" in result.columns
        else 1
    )
    result["_sort_open"] = (
        (~result["is_open"].fillna(False).astype(bool)).astype(int)
        if "is_open" in result.columns
        else 1
    )
    result["_sort_delay"] = (
        -pd.to_numeric(result["delay_days"], errors="coerce").fillna(0)
        if "delay_days" in result.columns
        else 0
    )
    result["_sort_created"] = pd.to_datetime(
        result["constraint_created_at"] if "constraint_created_at" in result.columns else None,
        errors="coerce",
    )
    result = result.sort_values(
        by=["_sort_block", "_sort_open", "_sort_delay", "_sort_created"],
        ascending=[True, True, True, False],
        kind="mergesort",
    )
    return result.drop(columns=[c for c in result.columns if c.startswith("_sort_")])


def _export_cell_value(row: pd.Series, field: str, kind: str) -> Any:
    """Typed cell value for Excel (not display-only strings for dates/money)."""
    if field == "work_package":
        raw = row.get("work_package")
        if raw is None or (isinstance(raw, float) and pd.isna(raw)) or not safe_str(raw):
            raw = row.get("iwp")
    elif field == "days_open_real":
        raw = row.get("days_open_real")
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            raw = row.get("days_open")
    elif field == "value_at_risk":
        raw = row.get("value_at_risk")
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            raw = row.get("plan_value")
    elif field == "deadline_status":
        raw = row.get("effective_deadline_status")
        if raw is None or (isinstance(raw, float) and pd.isna(raw)) or not safe_str(raw):
            raw = row.get("deadline_status")
    else:
        raw = row.get(field)

    if kind == "date":
        return safe_date(raw)
    if kind == "int":
        return int(safe_num(raw))
    if kind in {"money", "qty"}:
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            return None
        return safe_num(raw)
    if kind == "check":
        return display_check_status(raw)
    if kind == "resolution":
        return display_resolution_status(raw)
    if kind == "priority":
        text = safe_str(raw)
        return display_constraint_priority(text) if text else "—"
    if kind == "deadline":
        return display_deadline_status(raw) if safe_str(raw) else "—"
    if kind == "source":
        text = safe_str(raw)
        return display_deadline_source(text) if text else "—"
    if kind == "author":
        return display_author_empty(raw)
    if kind == "recorder":
        return display_recorder(raw)
    text = safe_str(raw)
    return text if text else "—"


def build_export_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Russian-labeled export frame; no internal/helper columns."""
    headers = [label for _, label, _ in EXPORT_COLUMN_SPECS]
    if df is None or df.empty:
        return pd.DataFrame(columns=headers)
    rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        item: dict[str, Any] = {}
        for field, label, kind in EXPORT_COLUMN_SPECS:
            item[label] = _export_cell_value(row, field, kind)
        rows.append(item)
    return pd.DataFrame(rows, columns=headers)


def resolve_export_source_df(
    *,
    full_df: pd.DataFrame,
    filtered_df: pd.DataFrame,
    scope: str,
    project: str,
) -> pd.DataFrame:
    """
    Export scope (independent of on-screen table when full-project).
    Filtered: current filtered dataframe.
    Full project: selected project, all months/disciplines; hide TEST projects.
    """
    if scope == EXPORT_SCOPE_FULL_PROJECT:
        if full_df is None or full_df.empty:
            return pd.DataFrame()
        out = full_df.copy()
        if "project_code" in out.columns:
            out = out[~out["project_code"].astype(str).isin(HIDDEN_PROJECT_CODES)]
            if project and project != ALL:
                out = out[out["project_code"].astype(str) == str(project)]
        return out.reset_index(drop=True)
    if filtered_df is None or filtered_df.empty:
        return pd.DataFrame()
    return filtered_df.copy().reset_index(drop=True)


def build_export_params_rows(
    *,
    formed_at: date,
    project: str,
    scope: str,
    row_count: int,
    month: str = ALL,
    discipline: str = ALL,
    queue: str = ALL,
    check_status: str = ALL,
    resolution_status: str = ALL,
    problem_owner: str = ALL,
    owner_name: str = ALL,
    coordinator: str = ALL,
    priority: str = ALL,
    deadline_status: str = ALL,
    open_mode: str = ALL,
) -> list[tuple[str, str]]:
    rows: list[tuple[str, str]] = [
        ("Дата формирования", formed_at.isoformat()),
        ("Проект", project or ALL),
        ("Объём выгрузки", scope),
    ]
    if scope == EXPORT_SCOPE_FILTERED:
        rows.extend(
            [
                (
                    "Месяц",
                    month if month != ALL else ALL,
                ),
                (
                    "Дисциплина",
                    discipline if discipline != ALL else ALL,
                ),
                ("Очередь", queue if queue != ALL else ALL),
                (
                    "Статус проверки",
                    format_check_filter_option(check_status)
                    if check_status != ALL
                    else ALL,
                ),
                (
                    "Статус устранения",
                    format_resolution_filter_option(resolution_status)
                    if resolution_status != ALL
                    else ALL,
                ),
                (
                    "Владелец ограничения",
                    problem_owner if problem_owner != ALL else ALL,
                ),
                (
                    "Исполнитель устранения",
                    owner_name if owner_name != ALL else ALL,
                ),
                (
                    "Координатор",
                    coordinator if coordinator != ALL else ALL,
                ),
                (
                    "Приоритет",
                    display_constraint_priority(priority)
                    if priority != ALL
                    else ALL,
                ),
                (
                    "Статус срока",
                    display_deadline_status(deadline_status)
                    if deadline_status != ALL
                    else ALL,
                ),
                ("Состояние реестра", open_mode if open_mode else ALL),
            ]
        )
    else:
        rows.extend(
            [
                ("Месяц", "все месяцы"),
                ("Дисциплина", "все дисциплины"),
            ]
        )
    rows.append(("Количество строк в выгрузке", str(int(row_count))))
    return rows


def safe_excel_filename(project: str, scope: str, month: str, formed_at: date) -> str:
    """Windows-safe .xlsx name."""
    proj = safe_str(project) or "проект"
    for ch in '<>:"/\\|?*':
        proj = proj.replace(ch, "_")
    proj = proj.replace(" ", "_")
    if scope == EXPORT_SCOPE_FULL_PROJECT:
        mid = "все_месяцы"
    else:
        mid = safe_str(month) or "месяц"
        for ch in '<>:"/\\|?*':
            mid = mid.replace(ch, "_")
        mid = mid.replace(" ", "_")
    return f"Реестр_ограничений_{proj}_{mid}_{formed_at.isoformat()}.xlsx"


def _export_param_value(
    params_rows: list[tuple[str, str]], key: str, default: str = "—"
) -> str:
    for k, v in params_rows:
        if k == key:
            text = safe_str(v)
            return text if text else default
    return default


def _compute_export_kpis(export_df: pd.DataFrame) -> list[tuple[str, Any]]:
    """Aggregates only from the export dataframe (no new business rules)."""
    total = int(len(export_df)) if export_df is not None else 0
    if export_df is None or export_df.empty:
        return [
            ("ВСЕГО ОГРАНИЧЕНИЙ", 0),
            ("ОТКРЫТЫХ", 0),
            ("HOLD", 0),
            ("FAIL", 0),
            ("WARNING", 0),
            ("ПРОСРОЧЕННЫХ", 0),
            ("СНЯТЫХ", 0),
            ("СТОИМОСТЬ BOQ ПОД ОГРАНИЧЕНИЯМИ", 0.0),
        ]

    check = (
        export_df["Статус проверки"].astype(str)
        if "Статус проверки" in export_df.columns
        else pd.Series("", index=export_df.index)
    )
    resolution = (
        export_df["Статус устранения"].astype(str)
        if "Статус устранения" in export_df.columns
        else pd.Series("", index=export_df.index)
    )
    deadline = (
        export_df["Статус срока"].astype(str)
        if "Статус срока" in export_df.columns
        else pd.Series("", index=export_df.index)
    )
    open_mask = resolution.isin(["Открыто", "В работе", "OPEN", "IN_PROGRESS"])
    cost_series = (
        pd.to_numeric(export_df["Стоимость под риском"], errors="coerce")
        if "Стоимость под риском" in export_df.columns
        else pd.Series(dtype=float)
    )
    cost_sum = float(cost_series.fillna(0).sum()) if not cost_series.empty else 0.0
    return [
        ("ВСЕГО ОГРАНИЧЕНИЙ", total),
        ("ОТКРЫТЫХ", int(open_mask.sum())),
        ("HOLD", int(check.isin(["Удержание", "HOLD"]).sum())),
        ("FAIL", int(check.isin(["Заблокировано", "FAIL"]).sum())),
        ("WARNING", int(check.isin(["Требует уточнения", "WARNING"]).sum())),
        (
            "ПРОСРОЧЕННЫХ",
            int(deadline.isin(["Просрочен", "Просрочено", "OVERDUE"]).sum()),
        ),
        ("СНЯТЫХ", int(resolution.isin(["Снято", "RESOLVED"]).sum())),
        ("СТОИМОСТЬ BOQ ПОД ОГРАНИЧЕНИЯМИ", cost_sum),
    ]


def _excel_fill(rgb: str):
    from openpyxl.styles import PatternFill

    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def _excel_font(
    *,
    bold: bool = False,
    size: int = 10,
    color: str = COLOR_NAVY,
    name: str = FONT_NAME,
):
    from openpyxl.styles import Font

    return Font(name=name, bold=bold, size=size, color=color)


def _excel_border():
    from openpyxl.styles import Border, Side

    thin = Side(style="thin", color=COLOR_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_cell_style(
    cell,
    *,
    fill=None,
    font=None,
    alignment=None,
    border=None,
    number_format: str | None = None,
) -> None:
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def _status_fill_for_value(value: Any):
    text = safe_str(value)
    if not text or text == "—":
        return None
    rgb = EXPORT_STATUS_FILLS.get(text)
    if rgb is None:
        rgb = EXPORT_STATUS_FILLS.get(text.upper())
    if rgb is None:
        return None
    return _excel_fill(rgb)


def build_registry_excel_bytes(
    export_df: pd.DataFrame,
    params_rows: list[tuple[str, str]],
) -> bytes:
    """In-memory .xlsx: registry + params sheets. Presentation styling only."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = EXPORT_SHEET_REGISTRY

    headers = list(export_df.columns) if not export_df.empty else [
        label for _, label, _ in EXPORT_COLUMN_SPECS
    ]
    ncol = max(1, len(headers))
    last_col_letter = get_column_letter(ncol)
    border = _excel_border()
    kind_by_header = {label: kind for _, label, kind in EXPORT_COLUMN_SPECS}

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_top = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right_top = Alignment(horizontal="right", vertical="top")
    wrap_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # --- A. Title band ---
    ws.merge_cells(start_row=EXPORT_ROW_TITLE, start_column=1, end_row=EXPORT_ROW_TITLE, end_column=ncol)
    title_cell = ws.cell(row=EXPORT_ROW_TITLE, column=1, value=EXPORT_TITLE)
    _apply_cell_style(
        title_cell,
        fill=_excel_fill(COLOR_NAVY),
        font=_excel_font(bold=True, size=14, color=COLOR_WHITE),
        alignment=center,
    )
    for col_idx in range(1, ncol + 1):
        cell = ws.cell(row=EXPORT_ROW_TITLE, column=col_idx)
        _apply_cell_style(cell, fill=_excel_fill(COLOR_NAVY), border=border)
    ws.row_dimensions[EXPORT_ROW_TITLE].height = 30

    # --- B. Subtitle ---
    ws.merge_cells(
        start_row=EXPORT_ROW_SUBTITLE,
        start_column=1,
        end_row=EXPORT_ROW_SUBTITLE,
        end_column=ncol,
    )
    sub_cell = ws.cell(row=EXPORT_ROW_SUBTITLE, column=1, value=EXPORT_SUBTITLE)
    _apply_cell_style(
        sub_cell,
        fill=_excel_fill(COLOR_LIGHT_BLUE),
        font=_excel_font(bold=False, size=10, color=COLOR_NAVY),
        alignment=center,
    )
    for col_idx in range(1, ncol + 1):
        _apply_cell_style(
            ws.cell(row=EXPORT_ROW_SUBTITLE, column=col_idx),
            fill=_excel_fill(COLOR_LIGHT_BLUE),
            border=border,
        )
    ws.row_dimensions[EXPORT_ROW_SUBTITLE].height = 28

    # --- Meta: project / date / scope (from params only) ---
    project = _export_param_value(params_rows, "Проект")
    formed = _export_param_value(params_rows, "Дата формирования")
    scope = _export_param_value(params_rows, "Объём выгрузки")
    meta_text = (
        f"Проект: {project}    |    "
        f"Дата формирования: {formed}    |    "
        f"Объём выгрузки: {scope}"
    )
    ws.merge_cells(
        start_row=EXPORT_ROW_META,
        start_column=1,
        end_row=EXPORT_ROW_META,
        end_column=ncol,
    )
    meta_cell = ws.cell(row=EXPORT_ROW_META, column=1, value=meta_text)
    _apply_cell_style(
        meta_cell,
        fill=_excel_fill(COLOR_META_BLUE),
        font=_excel_font(bold=True, size=9, color=COLOR_NAVY),
        alignment=center,
    )
    for col_idx in range(1, ncol + 1):
        _apply_cell_style(
            ws.cell(row=EXPORT_ROW_META, column=col_idx),
            fill=_excel_fill(COLOR_META_BLUE),
            border=border,
        )
    ws.row_dimensions[EXPORT_ROW_META].height = 22

    # --- C. KPI summary block ---
    kpis = _compute_export_kpis(export_df)
    kpi_slots = max(1, len(kpis))
    # Spread KPI pairs across available columns (min 1 col each).
    cols_per_kpi = max(1, ncol // kpi_slots)
    for i, (label, value) in enumerate(kpis):
        start_c = i * cols_per_kpi + 1
        end_c = start_c + cols_per_kpi - 1
        if i == len(kpis) - 1:
            end_c = ncol
        if start_c > ncol:
            break
        end_c = min(end_c, ncol)
        if end_c > start_c:
            ws.merge_cells(
                start_row=EXPORT_ROW_KPI_LABEL,
                start_column=start_c,
                end_row=EXPORT_ROW_KPI_LABEL,
                end_column=end_c,
            )
            ws.merge_cells(
                start_row=EXPORT_ROW_KPI_VALUE,
                start_column=start_c,
                end_row=EXPORT_ROW_KPI_VALUE,
                end_column=end_c,
            )
        label_cell = ws.cell(row=EXPORT_ROW_KPI_LABEL, column=start_c, value=label)
        value_cell = ws.cell(row=EXPORT_ROW_KPI_VALUE, column=start_c, value=value)
        kpi_fill = _excel_fill(COLOR_KPI_BG)
        _apply_cell_style(
            label_cell,
            fill=kpi_fill,
            font=_excel_font(bold=True, size=8, color=COLOR_NAVY),
            alignment=center,
            border=border,
        )
        is_money = label.startswith("СТОИМОСТЬ")
        _apply_cell_style(
            value_cell,
            fill=kpi_fill,
            font=_excel_font(bold=True, size=12, color=COLOR_NAVY),
            alignment=center,
            border=border,
            number_format='#,##0.00" ₽"' if is_money else "#,##0",
        )
        for col_idx in range(start_c, end_c + 1):
            for r in (EXPORT_ROW_KPI_LABEL, EXPORT_ROW_KPI_VALUE):
                c = ws.cell(row=r, column=col_idx)
                _apply_cell_style(c, fill=kpi_fill, border=border)
    ws.row_dimensions[EXPORT_ROW_KPI_LABEL].height = 18
    ws.row_dimensions[EXPORT_ROW_KPI_VALUE].height = 24

    # Spacer row 6
    ws.row_dimensions[6].height = 8

    # --- D. Table header ---
    header_fill = _excel_fill(COLOR_BLUE)
    header_font = _excel_font(bold=True, size=10, color=COLOR_WHITE)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=EXPORT_ROW_TABLE_HEADER, column=col_idx, value=header)
        _apply_cell_style(
            cell,
            fill=header_fill,
            font=header_font,
            alignment=center,
            border=border,
        )
    ws.row_dimensions[EXPORT_ROW_TABLE_HEADER].height = 48

    # --- E–I. Data rows ---
    status_headers = {
        "Статус проверки",
        "Статус устранения",
        "Приоритет",
        "Статус срока",
    }
    center_kinds = {"date", "check", "resolution", "priority", "deadline", "int"}
    for row_offset, (_, row) in enumerate(export_df.iterrows()):
        row_idx = EXPORT_ROW_DATA_START + row_offset
        alt = row_offset % 2 == 1
        row_fill = _excel_fill(COLOR_ALT_ROW) if alt else _excel_fill(COLOR_WHITE)
        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            kind = kind_by_header.get(header, "text")
            num_fmt = None
            if kind == "date" and isinstance(value, date):
                num_fmt = "DD.MM.YYYY"
            elif kind == "money" and value is not None:
                num_fmt = '#,##0.00" ₽"'
            elif kind == "qty" and value is not None:
                num_fmt = "#,##0.00"
            elif kind == "int" and value is not None:
                num_fmt = "#,##0"

            if header in EXPORT_WRAP_HEADERS:
                align = wrap_top
            elif kind == "money" or kind == "qty":
                align = right_top
            elif kind in center_kinds or header in {"BOQ", "Ед. изм.", "Очередь"}:
                align = center_top
            else:
                align = left_top

            fill = row_fill
            if header in status_headers:
                status_fill = _status_fill_for_value(value)
                if status_fill is not None:
                    fill = status_fill

            _apply_cell_style(
                cell,
                fill=fill,
                font=_excel_font(size=9, color=COLOR_NAVY),
                alignment=align,
                border=border,
                number_format=num_fmt,
            )
        ws.row_dimensions[row_idx].height = 22

    data_last_row = EXPORT_ROW_TABLE_HEADER
    if not export_df.empty:
        data_last_row = EXPORT_ROW_DATA_START + len(export_df) - 1

    # --- J. Column widths ---
    for col_idx, header in enumerate(headers, start=1):
        width = EXPORT_COLUMN_WIDTHS.get(header, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- L. Freeze panes: title/KPI/header + first two columns ---
    ws.freeze_panes = f"C{EXPORT_ROW_DATA_START}"

    # --- M. AutoFilter only on table ---
    if headers:
        ws.auto_filter.ref = (
            f"A{EXPORT_ROW_TABLE_HEADER}:{last_col_letter}{data_last_row}"
        )

    # --- N. Print / page setup ---
    ws.print_title_rows = f"{EXPORT_ROW_TABLE_HEADER}:{EXPORT_ROW_TABLE_HEADER}"
    ws.print_area = f"A1:{last_col_letter}{data_last_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- Parameters sheet (same corporate style) ---
    ws_params = wb.create_sheet(EXPORT_SHEET_PARAMS)
    ws_params.merge_cells("A1:B1")
    p_title = ws_params.cell(row=1, column=1, value="ПАРАМЕТРЫ ВЫГРУЗКИ")
    _apply_cell_style(
        p_title,
        fill=_excel_fill(COLOR_NAVY),
        font=_excel_font(bold=True, size=14, color=COLOR_WHITE),
        alignment=center,
    )
    _apply_cell_style(
        ws_params.cell(row=1, column=2),
        fill=_excel_fill(COLOR_NAVY),
        border=border,
    )
    ws_params.row_dimensions[1].height = 28

    ph_fill = _excel_fill(COLOR_BLUE)
    for col_idx, label in enumerate(("Параметр", "Значение"), start=1):
        cell = ws_params.cell(row=2, column=col_idx, value=label)
        _apply_cell_style(
            cell,
            fill=ph_fill,
            font=_excel_font(bold=True, size=10, color=COLOR_WHITE),
            alignment=center,
            border=border,
        )
    for idx, (key, value) in enumerate(params_rows, start=3):
        alt = (idx - 3) % 2 == 1
        fill = _excel_fill(COLOR_ALT_ROW) if alt else _excel_fill(COLOR_WHITE)
        c1 = ws_params.cell(row=idx, column=1, value=key)
        c2 = ws_params.cell(row=idx, column=2, value=value)
        _apply_cell_style(
            c1,
            fill=fill,
            font=_excel_font(bold=True, size=9, color=COLOR_NAVY),
            alignment=left_top,
            border=border,
        )
        _apply_cell_style(
            c2,
            fill=fill,
            font=_excel_font(size=9, color=COLOR_NAVY),
            alignment=left_top,
            border=border,
        )
    ws_params.column_dimensions["A"].width = 36
    ws_params.column_dimensions["B"].width = 52
    ws_params.freeze_panes = "A3"
    params_last = 2 + len(params_rows)
    ws_params.auto_filter.ref = f"A2:B{params_last}"
    ws_params.page_setup.orientation = "landscape"
    ws_params.page_setup.fitToPage = True
    ws_params.page_setup.fitToWidth = 1
    ws_params.page_setup.fitToHeight = 0
    ws_params.print_area = f"A1:B{params_last}"
    ws_params.sheet_properties.pageSetUpPr.fitToPage = True

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_display_table(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=[label for _, label, _ in TABLE_COLUMNS])

    rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        risk = row.get("value_at_risk")
        if risk is None or (isinstance(risk, float) and pd.isna(risk)):
            risk = row.get("plan_value")
        days_open_real = row.get("days_open_real")
        if days_open_real is None or (
            isinstance(days_open_real, float) and pd.isna(days_open_real)
        ):
            days_open_real = row.get("days_open")
        rows.append(
            {
                "Проект": display_dash(row.get("project_code")),
                "Месяц": display_dash(row.get("month_key")),
                "Дата возникновения": format_date_author_empty(
                    row.get("constraint_occurred_at")
                ),
                "Дата фиксации": format_date_ru(row.get("constraint_created_at")),
                "Очередь": display_dash(row.get("queue")),
                "Титул": display_dash(row.get("facility_building")),
                "Дисциплина": display_dash(row.get("construction_discipline")),
                "BOQ-код": display_dash(row.get("boq_code")),
                "Наименование работ": display_dash(row.get("boq_name")),
                "Пакет работ / IWP": display_dash(
                    row.get("work_package") or row.get("iwp")
                ),
                "Система": display_dash(row.get("system")),
                "Ед. изм.": display_dash(row.get("unit")),
                "Плановый объём": qty_ru(row.get("planned_qty")),
                "Плановая стоимость": money_ru(row.get("plan_value")),
                "Отдел": display_dash(row.get("responsible_department")),
                "Статус проверки": display_check_status_table(row.get("check_status")),
                "Статус устранения": display_resolution_status_table(
                    row.get("resolution_status")
                ),
                "Тип ограничения": display_dash(row.get("constraint_category")),
                "Приоритет": display_priority_table(row.get("constraint_priority")),
                "Суть ограничения": display_dash(row.get("problem_summary")),
                "Владелец ограничения": display_author_empty(row.get("problem_owner")),
                "Исполнитель": display_author_empty(row.get("owner_name")),
                "Координатор": display_author_empty(
                    row.get("subcontractor_coordinator")
                ),
                "Влияние": display_author_empty(row.get("problem_impact")),
                "Требуемое действие": display_author_empty(row.get("required_action")),
                "Статус срока": display_deadline_status(
                    row.get("effective_deadline_status") or row.get("deadline_status")
                ),
                "Плановая дата устранения": format_date_author_empty(
                    row.get("target_resolution_date")
                ),
                "Следующий контроль": format_date_author_empty(
                    row.get("next_control_date")
                ),
                "Фактическая дата устранения": format_date_ru(
                    row.get("actual_resolution_date")
                ),
                "Дней открыто": int(safe_num(days_open_real)),
                "Задержка, дней": int(safe_num(row.get("delay_days"))),
                "Стоимость под риском": money_ru(risk),
                "Примечание": truncate_comment(row.get("comment")),
                "Зафиксировал ограничение": display_recorder(row.get("created_by")),
                "Последнее изменение": display_updated_by(row.get("updated_by")),
                "Кто снял": display_resolved_by(row.get("resolved_by")),
                "Дата последнего действия": format_date_ru(row.get("last_action_at")),
            }
        )
    return pd.DataFrame(rows)


def build_column_config() -> dict[str, Any]:
    config: dict[str, Any] = {}
    wide = {
        "Наименование работ",
        "Суть ограничения",
        "Влияние",
        "Требуемое действие",
        "Примечание",
        "Титул",
        "Пакет работ / IWP",
        "Дата последнего действия",
    }
    for _, label, width in TABLE_COLUMNS:
        if label in {"Задержка, дней", "Дней открыто"}:
            config[label] = st.column_config.NumberColumn(label, width=width, format="%d")
        elif label == "Дата последнего действия":
            config[label] = st.column_config.TextColumn(
                "Дата последнего действия",
                width=width,
                help="Дата последнего действия по ограничению",
            )
        elif label == "Примечание":
            config[label] = st.column_config.TextColumn(
                label,
                width=width,
                help="В таблице — сокращённый текст; полный комментарий в карточке ниже",
            )
        else:
            config[label] = st.column_config.TextColumn(
                label,
                width=width if label in wide or width >= 180 else width,
            )
    return config


# ---------------------------------------------------------------------------
# Page
# ---------------------------------------------------------------------------

st.title("Реестр ограничений производства работ")
st.caption(
    "Единый реестр ограничений, препятствующих допуску, началу или продолжению работ месячного плана."
)

try:
    with st.spinner("Загрузка реестра ограничений…"):
        full_df = load_constraint_registry()
except Exception as exc:  # noqa: BLE001
    st.error(f"Не удалось загрузить реестр ограничений: {exc}")
    st.stop()

if full_df is None or full_df.empty:
    st.info("В реестре пока нет ограничений.")
    st.stop()

# UI-only: hide fixture projects from the registry view (DB unchanged).
if "project_code" in full_df.columns:
    full_df = full_df[
        ~full_df["project_code"].astype(str).isin(HIDDEN_PROJECT_CODES)
    ].copy()
if full_df.empty:
    st.info("В реестре пока нет ограничений.")
    st.stop()

projects = [
    p
    for p in (
        unique_sorted(full_df["project_code"])
        if "project_code" in full_df.columns
        else []
    )
    if p not in HIDDEN_PROJECT_CODES
]
month_options = list(MONTH_OPTIONS_2026)

default_project = projects[0] if projects else ALL
default_month = (
    "август-2026"
    if "август-2026" in month_options
    else (month_options[0] if month_options else ALL)
)

if "reg_project" not in st.session_state:
    st.session_state["reg_project"] = default_project
if "reg_month" not in st.session_state:
    st.session_state["reg_month"] = default_month
if "reg_open_mode" not in st.session_state:
    st.session_state["reg_open_mode"] = OPEN_ONLY

# Keep selectbox values valid after data changes
if st.session_state["reg_project"] not in projects and projects:
    st.session_state["reg_project"] = default_project
# Drop test / non-calendar months from session (e.g. test-r2-ui-2026)
if st.session_state.get("reg_month") not in month_options:
    st.session_state["reg_month"] = default_month
# Migrate legacy open-mode labels
_open_mode_options = [MODE_ALL, OPEN_ONLY]
if st.session_state.get("reg_open_mode") not in _open_mode_options:
    legacy = safe_str(st.session_state.get("reg_open_mode"))
    if legacy in {"Открытые", "OPEN_ONLY", OPEN_ONLY}:
        st.session_state["reg_open_mode"] = OPEN_ONLY
    else:
        st.session_state["reg_open_mode"] = MODE_ALL

f1, f2, f3, f4 = st.columns(4)
with f1:
    project = st.selectbox("Проект", options=projects or [ALL], key="reg_project")
with f2:
    month = st.selectbox("Месяц", options=month_options, key="reg_month")
with f3:
    open_mode = st.selectbox(
        "Состояние реестра",
        options=_open_mode_options,
        key="reg_open_mode",
    )
with f4:
    search = st.text_input(
        "Поиск по BOQ-коду или наименованию",
        value="",
        key="reg_search",
    )

# Scope for dependent filter options: project + month only (no cascade wipe).
scoped = full_df
if project != ALL and "project_code" in scoped.columns:
    scoped = scoped[scoped["project_code"].astype(str) == project]
if month != ALL and "month_key" in scoped.columns:
    scoped = scoped[scoped["month_key"].astype(str) == month]

facilities = [ALL] + (
    unique_sorted(scoped["facility_building"])
    if "facility_building" in scoped.columns
    else []
)
# Discipline options from project+month scope (+ canonical pipe testing always).
_discipline_values = (
    unique_sorted(scoped["construction_discipline"])
    if "construction_discipline" in scoped.columns
    else []
)
disciplines = [ALL] + sorted(
    {*_discipline_values, PIPE_TESTING_DISCIPLINE},
    key=lambda v: v.casefold(),
)
departments = [ALL] + (
    unique_sorted(scoped["responsible_department"])
    if "responsible_department" in scoped.columns
    else []
)
queues = [ALL] + (
    unique_sorted(scoped["queue"]) if "queue" in scoped.columns else []
)
owners = [ALL] + (
    unique_sorted(scoped["problem_owner"]) if "problem_owner" in scoped.columns else []
)
executors = [ALL] + (
    unique_sorted(scoped["owner_name"]) if "owner_name" in scoped.columns else []
)
coordinators = [ALL] + (
    unique_sorted(scoped["subcontractor_coordinator"])
    if "subcontractor_coordinator" in scoped.columns
    else []
)

# Fixed enum options always available (not cascade-shrunk by other filters).
check_status_options = [ALL] + CHECK_STATUS_OPTIONS
resolution_status_options = [ALL] + RESOLUTION_STATUS_OPTIONS
priority_options = [ALL] + PRIORITY_FILTER_OPTIONS
deadline_status_options = [ALL] + DEADLINE_STATUS_FILTER_OPTIONS

for key, options in (
    ("reg_facility", facilities),
    ("reg_discipline", disciplines),
    ("reg_department", departments),
    ("reg_queue", queues),
    ("reg_problem_owner", owners),
    ("reg_owner_name", executors),
    ("reg_coordinator", coordinators),
    ("reg_priority", priority_options),
    ("reg_eff_deadline", deadline_status_options),
    ("reg_deadline_overdue", [ALL, "Да", "Нет"]),
    ("reg_control_overdue", [ALL, "Да", "Нет"]),
    ("reg_check_status", check_status_options),
    ("reg_resolution_status", resolution_status_options),
):
    if st.session_state.get(key) not in options:
        st.session_state[key] = ALL

f5, f6, f7, f8 = st.columns(4)
with f5:
    facility = st.selectbox("Титул", options=facilities, key="reg_facility")
with f6:
    discipline = st.selectbox("Дисциплина", options=disciplines, key="reg_discipline")
with f7:
    department = st.selectbox("Отдел", options=departments, key="reg_department")
with f8:
    queue_f = st.selectbox("Очередь", options=queues, key="reg_queue")

f9, f10 = st.columns(2)
with f9:
    check_status = st.selectbox(
        "Статус проверки",
        options=check_status_options,
        format_func=format_check_filter_option,
        key="reg_check_status",
    )
with f10:
    resolution_status = st.selectbox(
        "Статус устранения",
        options=resolution_status_options,
        format_func=format_resolution_filter_option,
        key="reg_resolution_status",
        help=(
            "Открыто=OPEN, В работе=IN_PROGRESS, Снято=RESOLVED, Отменено=CANCELLED. "
            "Работает вместе с «Состояние реестра»."
        ),
    )

f11, f12, f13, f14 = st.columns(4)
with f11:
    problem_owner_f = st.selectbox(
        "Владелец ограничения", options=owners, key="reg_problem_owner"
    )
with f12:
    owner_name_f = st.selectbox(
        "Исполнитель устранения", options=executors, key="reg_owner_name"
    )
with f13:
    coordinator_f = st.selectbox(
        "Координатор от субподрядчика", options=coordinators, key="reg_coordinator"
    )
with f14:
    priority_f = st.selectbox(
        "Приоритет",
        options=priority_options,
        format_func=lambda v: ALL if v == ALL else display_constraint_priority(v),
        key="reg_priority",
    )

f15, f16, f17 = st.columns(3)
with f15:
    eff_deadline_f = st.selectbox(
        "Статус срока",
        options=deadline_status_options,
        format_func=lambda v: ALL if v == ALL else display_deadline_status(v),
        key="reg_eff_deadline",
    )
with f16:
    deadline_overdue_f = st.selectbox(
        "Просрочен срок",
        options=[ALL, "Да", "Нет"],
        key="reg_deadline_overdue",
    )
with f17:
    control_overdue_f = st.selectbox(
        "Просрочен следующий контроль",
        options=[ALL, "Да", "Нет"],
        key="reg_control_overdue",
    )

filtered = apply_filters(
    full_df,
    project=project,
    month=month,
    facility=facility,
    discipline=discipline,
    department=department,
    check_status=check_status,
    resolution_status=resolution_status,
    open_mode=open_mode,
    search=search,
    problem_owner=problem_owner_f,
    owner_name=owner_name_f,
    subcontractor_coordinator=coordinator_f,
    constraint_priority=priority_f,
    effective_deadline_status=eff_deadline_f,
    is_deadline_overdue=deadline_overdue_f,
    is_next_control_overdue=control_overdue_f,
    queue=queue_f,
)
filtered = sort_registry_rows(filtered)
kpi = build_registry_summary(filtered)

# Success banner from previous resolve (before rerun reload)
success_payload = st.session_state.pop("reg_resolve_success", None)
if isinstance(success_payload, dict):
    st.success("Ограничение снято")
    new_check = display_check_status(success_payload.get("new_check_status"))
    new_res = display_resolution_status(success_payload.get("new_resolution_status"))
    line_summary = success_payload.get("line_summary") or {}
    if not isinstance(line_summary, dict):
        line_summary = {}
    outcome = safe_str(line_summary.get("admission_outcome") or success_payload.get("admission_outcome")).upper()
    blockers = success_payload.get("remaining_blockers") or []
    if not isinstance(blockers, list):
        blockers = []
    st.write(
        f"Новый статус ограничения: **{new_check}** / **{new_res}**"
    )
    st.write(
        f"Итог по BOQ (admission_outcome): **{outcome or '—'}** · "
        f"оставшихся blockers: **{len(blockers)}**"
    )
    outcome_msg = ADMISSION_OUTCOME_MESSAGES.get(outcome)
    if outcome_msg:
        st.info(outcome_msg)
    st.caption("BOQ не добавляется в паспорт автоматически — только доступен для контролируемого включения.")

resolve_error = st.session_state.pop("reg_resolve_error", None)
if resolve_error:
    st.error(str(resolve_error))

update_success = st.session_state.pop("reg_update_success", None)
if isinstance(update_success, dict):
    st.success("Ограничение обновлено")
    changed = update_success.get("changed_fields") or []
    if isinstance(changed, list) and changed:
        st.caption("Изменённые поля: " + ", ".join(str(x) for x in changed))
update_error = st.session_state.pop("reg_update_error", None)
if update_error:
    st.error(str(update_error))

st.markdown("### Показатели")

st.caption("Слой 1 · Ограничения (по constraint_id)")
k1, k2, k3, k4, k5, k6, k7 = st.columns(7)
k1.metric("Всего ограничений", kpi["constraints_total"])
k2.metric("Открытых", kpi["constraints_open"])
k3.metric("HOLD", kpi["constraints_hold"])
k4.metric("FAIL", kpi["constraints_fail"])
k5.metric("WARNING", kpi["constraints_warning"])
k6.metric("Просроченных", kpi["constraints_overdue"])
k7.metric("Снятых", kpi["constraints_resolved"])

st.caption("Слой 2 · BOQ под ограничениями (уникальный plan_line_id)")
b1, b2, b3, b4, b5, b6 = st.columns(6)
b1.metric("BOQ под ограничением", kpi["boq_under_constraint_count"])
b2.metric("Стоимость BOQ под риском", money_ru(kpi["boq_cost_at_risk"]))
b3.metric("Уникальных титулов", kpi["unique_facility_count"])
b4.metric("Уникальных дисциплин", kpi["unique_discipline_count"])
b5.metric("Уникальных систем", kpi["unique_system_count"])
b6.metric("Уникальных IWP", kpi["unique_iwp_count"])

with st.expander("Экспорт реестра", expanded=False):
    export_scope = st.radio(
        "Объём выгрузки",
        options=EXPORT_SCOPE_OPTIONS,
        index=0,
        key="reg_export_scope",
    )
    export_source = resolve_export_source_df(
        full_df=full_df,
        filtered_df=filtered,
        scope=export_scope,
        project=project,
    )
    export_formed = date.today()
    if export_source.empty:
        st.info("По выбранным фильтрам нет данных для выгрузки.")
    else:
        export_df = build_export_dataframe(export_source)
        params_rows = build_export_params_rows(
            formed_at=export_formed,
            project=project,
            scope=export_scope,
            row_count=len(export_df),
            month=month,
            discipline=discipline,
            queue=queue_f,
            check_status=check_status,
            resolution_status=resolution_status,
            problem_owner=problem_owner_f,
            owner_name=owner_name_f,
            coordinator=coordinator_f,
            priority=priority_f,
            deadline_status=eff_deadline_f,
            open_mode=open_mode,
        )
        xlsx_bytes = build_registry_excel_bytes(export_df, params_rows)
        file_name = safe_excel_filename(
            project, export_scope, month, export_formed
        )
        st.caption(f"Строк к выгрузке: {len(export_df)}")
        st.download_button(
            "Скачать Excel",
            data=xlsx_bytes,
            file_name=file_name,
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            key="reg_export_download",
        )

st.markdown("### Реестр")
st.caption(
    f"Строк в таблице: {len(filtered)}. "
    "Выберите строку для карточки ограничения."
)

selected_constraint_id: Optional[str] = None

if filtered.empty:
    st.info("По выбранным фильтрам ограничений нет.")
else:
    display_df = build_display_table(filtered)
    # Keep stable positional index for selection → constraint_id mapping
    display_df = display_df.reset_index(drop=True)
    filtered_indexed = filtered.reset_index(drop=True)
    row_constraint_ids = [
        safe_str(cid) for cid in filtered_indexed["constraint_id"].tolist()
    ]

    st.dataframe(
        style_registry_display(display_df),
        use_container_width=True,
        hide_index=True,
        height=REGISTRY_TABLE_HEIGHT_PX,
        on_select="rerun",
        selection_mode="single-row",
        key=REGISTRY_SELECT_KEY,
        column_config=build_column_config(),
    )
    selected_constraint_id = resolve_registry_selection(row_constraint_ids)

# Resolve selected row from full_df (fresh fields even if filtered mode changes)
selected_row: Optional[pd.Series] = None
if selected_constraint_id:
    matches = full_df[full_df["constraint_id"].astype(str) == selected_constraint_id]
    if not matches.empty:
        selected_row = matches.iloc[0]
    else:
        st.warning("Выбранное ограничение не найдено в текущей загрузке реестра.")
        selected_constraint_id = None

if selected_row is not None and selected_constraint_id:
    with st.expander("Карточка ограничения", expanded=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f"**constraint_id:** `{selected_constraint_id}`")
            st.markdown(f"**Проект:** {display_dash(selected_row.get('project_code'))}")
            st.markdown(f"**Месяц:** {display_dash(selected_row.get('month_key'))}")
            st.markdown(f"**Титул:** {display_dash(selected_row.get('facility_building'))}")
            st.markdown(
                f"**Дисциплина:** {display_dash(selected_row.get('construction_discipline'))}"
            )
            st.markdown(f"**BOQ-код:** {display_dash(selected_row.get('boq_code'))}")
            st.markdown(
                f"**Наименование работ:** {display_dash(selected_row.get('boq_name'))}"
            )
            st.markdown(
                f"**Дата возникновения:** {format_date_author_empty(selected_row.get('constraint_occurred_at'))}"
            )
            st.markdown(
                f"**Дата регистрации:** {format_date_ru(selected_row.get('constraint_created_at'))}"
            )
            days_open_real = selected_row.get("days_open_real")
            if days_open_real is None or (
                isinstance(days_open_real, float) and pd.isna(days_open_real)
            ):
                days_open_real = selected_row.get("days_open")
            st.markdown(f"**Дней открыто (real):** {int(safe_num(days_open_real))}")
        with c2:
            st.markdown(
                f"**Отдел:** {display_dash(selected_row.get('responsible_department'))}"
            )
            st.markdown(f"**Проверка:** {display_dash(selected_row.get('check_name'))}")
            st.markdown(
                f"**Статус проверки:** {display_check_status(selected_row.get('check_status'))}"
            )
            st.markdown(
                f"**Статус устранения:** {display_resolution_status(selected_row.get('resolution_status'))}"
            )
            st.markdown(
                f"**Категория:** {display_author_empty(selected_row.get('constraint_category'))}"
            )
            st.markdown(
                f"**Приоритет:** {display_constraint_priority(selected_row.get('constraint_priority'))}"
            )
            st.markdown(
                f"**Владелец ограничения:** {display_author_empty(selected_row.get('problem_owner'))}"
            )
            st.markdown(
                f"**Исполнитель устранения:** {display_author_empty(selected_row.get('owner_name'))}"
            )
            st.markdown(
                f"**Координатор от субподрядчика:** {display_author_empty(selected_row.get('subcontractor_coordinator'))}"
            )
            st.markdown(
                f"**Зафиксировал ограничение:** {display_recorder(selected_row.get('created_by'))}"
            )
            st.markdown(
                f"**Последнее изменение:** {display_updated_by(selected_row.get('updated_by'))}"
            )
            st.markdown(
                f"**Снял:** {display_resolved_by(selected_row.get('resolved_by'))}"
            )
        with c3:
            st.markdown(
                f"**Суть ограничения:** {display_author_empty(selected_row.get('problem_summary'))}"
            )
            st.markdown(
                f"**Подробное описание проблемы:** {display_author_empty(selected_row.get('problem_description') or selected_row.get('problem_summary'))}"
            )
            st.markdown(
                f"**Влияние:** {display_author_empty(selected_row.get('problem_impact'))}"
            )
            st.markdown(
                f"**Требуемое действие:** {display_author_empty(selected_row.get('required_action'))}"
            )
            st.markdown(
                f"**Статус срока:** {display_deadline_status(selected_row.get('effective_deadline_status') or selected_row.get('deadline_status'))}"
            )
            st.markdown(
                f"**Источник срока:** {display_deadline_source(selected_row.get('deadline_source'))}"
            )
            st.markdown(
                f"**Плановая дата устранения:** {format_date_author_empty(selected_row.get('target_resolution_date'))}"
            )
            st.markdown(
                f"**Дата следующего контроля:** {format_date_author_empty(selected_row.get('next_control_date'))}"
            )
            overdue_deadline = bool(selected_row.get("is_deadline_overdue"))
            overdue_control = bool(selected_row.get("is_next_control_overdue"))
            st.markdown(
                f"**Просрочен срок:** {'Да' if overdue_deadline else 'Нет'}"
                f" · **Просрочен контроль:** {'Да' if overdue_control else 'Нет'}"
            )
            risk = selected_row.get("value_at_risk")
            if risk is None or (isinstance(risk, float) and pd.isna(risk)):
                risk = selected_row.get("plan_value")
            st.markdown(f"**Стоимость под риском:** {money_ru(risk)}")

        st.markdown("**Последний комментарий**")
        st.text(display_dash(selected_row.get("comment")))

    resolution_now = safe_str(selected_row.get("resolution_status")).upper()
    can_update = can_edit_update_form(resolution_now)
    can_resolve = resolution_now not in CLOSED_RESOLUTIONS

    # ---- R2: Скорректировать ограничение ----
    with st.expander("Скорректировать ограничение", expanded=False):
        baseline = row_to_update_baseline(selected_row)
        hydrate_update_form_state(st.session_state, selected_constraint_id, baseline)
        baseline = dict(st.session_state.get("reg_upd_baseline") or baseline)

        if not can_update:
            st.info(
                f"Обновление недоступно для статуса {resolution_now}. "
                "Форма доступна только для OPEN / IN_PROGRESS."
            )
        else:
            # Date modes/inputs MUST stay outside st.form so mode changes rerun
            # immediately and date_input appears without an intermediate submit.
            date_modes_ui, date_values_ui = render_update_date_controls()

            with st.form("reg_update_form", clear_on_submit=False):
                t1, t2 = st.columns(2)
                with t1:
                    problem_owner_v = st.text_input(
                        "Владелец ограничения", key="reg_upd_problem_owner"
                    )
                    owner_name_v = st.text_input(
                        "Исполнитель устранения", key="reg_upd_owner_name"
                    )
                    coordinator_v = st.text_input(
                        "Координатор от субподрядчика",
                        key="reg_upd_subcontractor_coordinator",
                    )
                    category_v = st.text_input(
                        "Категория ограничения", key="reg_upd_constraint_category"
                    )
                with t2:
                    priority_opts = [""] + list(CONSTRAINT_PRIORITY_DISPLAY.keys())
                    priority_v = st.selectbox(
                        "Приоритет",
                        options=priority_opts,
                        format_func=lambda v: "—" if not v else display_constraint_priority(v),
                        key="reg_upd_constraint_priority",
                    )
                    dl_status_opts = [""] + list(DEADLINE_STATUS_DISPLAY.keys())
                    deadline_status_v = st.selectbox(
                        "Статус срока",
                        options=dl_status_opts,
                        format_func=lambda v: "—" if not v else display_deadline_status(v),
                        key="reg_upd_deadline_status",
                    )
                    dl_source_opts = [""] + list(DEADLINE_SOURCE_DISPLAY.keys())
                    deadline_source_v = st.selectbox(
                        "Источник срока",
                        options=dl_source_opts,
                        format_func=lambda v: "—" if not v else display_deadline_source(v),
                        key="reg_upd_deadline_source",
                    )

                problem_description_v = st.text_area(
                    "Подробное описание проблемы",
                    height=100,
                    key="reg_upd_problem_description",
                )
                problem_impact_v = st.text_area(
                    "Влияние",
                    height=80,
                    key="reg_upd_problem_impact",
                    help=(
                        "На что влияет ограничение: начало/продолжение работ, закупка, "
                        "мобилизация, производительность, срок, стоимость, приёмка, ПНР и др."
                    ),
                )
                required_action_v = st.text_area(
                    "Требуемое действие",
                    height=80,
                    key="reg_upd_required_action",
                )

                updated_by_v = st.text_input("Кто обновил", key="reg_upd_updated_by")
                update_comment_v = st.text_area(
                    "Комментарий обновления",
                    height=80,
                    key="reg_upd_update_comment",
                )

                current_preview = collect_update_form_current(
                    text_values={
                        "problem_owner": problem_owner_v,
                        "owner_name": owner_name_v,
                        "subcontractor_coordinator": coordinator_v,
                        "constraint_category": category_v,
                        "problem_description": problem_description_v,
                        "problem_impact": problem_impact_v,
                        "required_action": required_action_v,
                    },
                    enum_values={
                        "constraint_priority": priority_v,
                        "deadline_status": deadline_status_v,
                        "deadline_source": deadline_source_v,
                    },
                    date_modes=date_modes_ui,
                    date_values=date_values_ui,
                )
                # Dates with KEEP are omitted from current; for caption, merge baseline for KEEP
                preview_for_patch = dict(current_preview)
                dirty_preview = build_update_dirty_patch(baseline, preview_for_patch)
                st.caption(f"Будет изменено полей: {len(dirty_preview)}")

                submitted_update = st.form_submit_button(
                    "Сохранить обновление",
                    type="primary",
                )

            if submitted_update:
                current_vals = collect_update_form_current(
                    text_values={
                        "problem_owner": problem_owner_v,
                        "owner_name": owner_name_v,
                        "subcontractor_coordinator": coordinator_v,
                        "constraint_category": category_v,
                        "problem_description": problem_description_v,
                        "problem_impact": problem_impact_v,
                        "required_action": required_action_v,
                    },
                    enum_values={
                        "constraint_priority": priority_v,
                        "deadline_status": deadline_status_v,
                        "deadline_source": deadline_source_v,
                    },
                    date_modes=date_modes_ui,
                    date_values=date_values_ui,
                )
                dirty = build_update_dirty_patch(baseline, current_vals)
                if not safe_str(updated_by_v):
                    st.error("Укажите, кто обновил ограничение.")
                elif not safe_str(update_comment_v):
                    st.error("Укажите комментарий обновления.")
                elif not dirty:
                    st.info("Нет изменений для сохранения.")
                else:
                    try:
                        normalized = normalize_update_patch(dirty)
                    except ValueError as exc:
                        st.error(str(exc))
                        normalized = None
                    if normalized is not None:
                        with st.spinner("Обновление ограничения…"):
                            result = update_constraint(
                                constraint_id=selected_constraint_id,
                                updated_by=updated_by_v,
                                update_comment=update_comment_v,
                                patch=normalized,
                            )
                        if result.get("ok") and result.get("status") == "no_changes":
                            st.info("Нет изменений для сохранения.")
                        elif result.get("ok"):
                            clear_registry_read_caches()
                            st.session_state["reg_update_success"] = result.get("data") or {
                                "changed_fields": list(normalized.keys())
                            }
                            st.session_state["reg_upd_update_comment"] = ""
                            st.rerun()
                        else:
                            st.session_state["reg_update_error"] = (
                                result.get("error") or "Не удалось обновить ограничение"
                            )
                            # Keep form values; show error without clearing widgets
                            st.error(st.session_state["reg_update_error"])

    with st.expander("Снять ограничение", expanded=False):
        if can_resolve:
            confirm = st.checkbox(
                "Подтверждаю, что причина ограничения фактически устранена",
                value=False,
                key="reg_resolve_confirm",
            )
            with st.form("reg_resolve_form", clear_on_submit=False):
                actual_date = st.date_input(
                    "Фактическая дата устранения",
                    value=date.today(),
                    key="reg_resolve_actual_date",
                )
                closed_by = st.text_input(
                    "Кто закрывает",
                    value="",
                    key="reg_resolve_closed_by",
                )
                resolution_comment = st.text_area(
                    "Комментарий об устранении",
                    value="",
                    height=120,
                    key="reg_resolve_comment",
                )
                submitted = st.form_submit_button(
                    "Подтвердить снятие ограничения",
                    type="primary",
                    disabled=not confirm,
                )

            if submitted:
                if not confirm:
                    st.warning("Нужно подтверждение снятия ограничения.")
                elif not safe_str(closed_by):
                    st.error("Укажите, кто закрывает ограничение.")
                elif not safe_str(resolution_comment):
                    st.error("Укажите комментарий об устранении.")
                elif actual_date is None:
                    st.error("Укажите фактическую дату устранения.")
                else:
                    with st.spinner("Снятие ограничения…"):
                        result = resolve_constraint(
                            constraint_id=selected_constraint_id,
                            actual_resolution_date=actual_date,
                            resolution_comment=resolution_comment,
                            closed_by=closed_by,
                            evidence_payload=None,
                        )
                    # Extra clear for page 21/23 helpers if already loaded
                    clear_constraint_registry_caches()
                    if result.get("ok"):
                        st.session_state["reg_resolve_success"] = result.get("data") or {}
                        st.session_state.pop("reg_selected_constraint_id", None)
                        st.session_state.pop("reg_resolve_confirm", None)
                        st.rerun()
                    else:
                        st.error(result.get("error") or "Не удалось снять ограничение")
        else:
            st.info(
                f"Ограничение уже закрыто ({resolution_now}). Форма снятия недоступна."
            )

    with st.expander("История событий ограничения", expanded=False):
        st.markdown(
            f"**Зафиксировал:** {display_recorder(selected_row.get('created_by'))}"
        )
        st.markdown(
            f"**Последнее изменение:** {display_updated_by(selected_row.get('updated_by'))}"
        )
        st.markdown(
            f"**Снял:** {display_resolved_by(selected_row.get('resolved_by'))}"
        )
        st.divider()
        try:
            events_df = load_constraint_events(selected_constraint_id)
        except Exception as exc:  # noqa: BLE001
            st.warning(f"Не удалось загрузить историю событий: {exc}")
            events_df = pd.DataFrame()

        if events_df is None or events_df.empty:
            st.caption("История событий пока отсутствует")
        else:
            for _, ev in events_df.iterrows():
                ev_type = safe_str(ev.get("event_type")).upper()
                when = format_datetime_ru(ev.get("performed_at"))
                who = display_dash(ev.get("performed_by"))
                comment = display_dash(ev.get("event_comment"))
                if ev_type == "UPDATED":
                    parsed = parse_constraint_event_payload(ev.get("event_payload"))
                    st.markdown(f"**UPDATED** · {when} · {who}")
                    st.caption(f"Комментарий: {comment}")
                    changed = parsed.get("changed_fields") or []
                    old_values = parsed.get("old_values") or {}
                    new_values = parsed.get("new_values") or {}
                    if changed:
                        lines = []
                        for field in changed:
                            label = FIELD_LABELS_RU.get(field, field)
                            old_v = format_update_value_display(
                                field, old_values.get(field)
                            )
                            new_v = format_update_value_display(
                                field, new_values.get(field)
                            )
                            lines.append(f"- {label}: {old_v} → {new_v}")
                        st.markdown("\n".join(lines))
                    else:
                        st.caption("changed_fields пуст")
                else:
                    st.markdown(
                        f"**{ev_type or 'EVENT'}** · {when} · {who} · "
                        f"{display_check_status(ev.get('old_check_status'))}→"
                        f"{display_check_status(ev.get('new_check_status'))} · "
                        f"{display_resolution_status(ev.get('old_resolution_status'))}→"
                        f"{display_resolution_status(ev.get('new_resolution_status'))}"
                    )
                    st.caption(f"Комментарий: {comment}")
                st.divider()
else:
    st.markdown("### Карточка ограничения")
    st.info(
        "Выберите строку в реестре, чтобы открыть карточку, обновление и снятие."
    )
