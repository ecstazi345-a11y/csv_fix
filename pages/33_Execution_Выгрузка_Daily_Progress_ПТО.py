import re
from datetime import datetime, timezone
from io import BytesIO

import streamlit as st
import pandas as pd

from services.supabase_client import supabase

st.set_page_config(layout="wide")

st.title("Выгрузка Daily Progress для ПТО")
st.caption(
    "Страница предназначена для быстрой выгрузки Daily Progress в контур ПТО для проверки объёмов, "
    "КС-6а и исполнительной документации."
)

EXPORT_COLUMNS = [
    "Work_Date",
    "Month_Key",
    "Facility_Building",
    "Construction_Discipline",
    "Foreman",
    "Crew_ID",
    "boq_name_clean",
    "IWP_ID_clean",
    "system_label_clean",
    "Unit_of_Measure",
    "Quantity_Today",
    "Unit_Rate",
    "EV_DAY_VALUE",
    "validation_status_clean",
]

XLSX_TEXT_COLUMNS = {
    "Month_Key",
    "Facility_Building",
    "Construction_Discipline",
    "Foreman",
    "Crew_ID",
    "boq_name_clean",
    "IWP_ID_clean",
    "system_label_clean",
    "Unit_of_Measure",
    "validation_status_clean",
}

XLSX_NUMERIC_COLUMNS = {
    "Quantity_Today",
    "Unit_Rate",
    "EV_DAY_VALUE",
}

XLSX_DATE_COLUMNS = {"Work_Date"}

XLSX_NUM_FORMATS = {
    "Quantity_Today": "0.00",
    "Unit_Rate": "#,##0.00",
    "EV_DAY_VALUE": "#,##0.00",
}

WORK_DATE_FALLBACK_COLUMNS = [
    "work_date",
    "Work_Date",
    "created_time",
    "created_at",
]

XLSX_VISIBLE_TOTAL_ROW = 7
XLSX_DATA_HEADER_ROW = 8
XLSX_DATA_START_ROW = 9
XLSX_SNAPSHOT_START_COL = 4

XLSX_TITLE_FILL = "1F3864"
XLSX_HEADER_FILL = "2F4F6F"
XLSX_ZEBRA_FILL = "F5F7FA"
XLSX_ZERO_QTY_FILL = "ECECEC"
XLSX_TOTAL_FILL = "D9E2F3"

EXPORT_FIELD_MAP = {
    "Work_Date": ["work_date", "Work_Date", "date", "work_day"],
    "Month_Key": ["month_key", "Month_Key", "month"],
    "Facility_Building": ["facility_building", "Facility_Building", "title", "Title"],
    "Construction_Discipline": ["construction_discipline", "Construction_Discipline", "discipline", "Discipline"],
    "Foreman": ["foreman", "Foreman"],
    "Crew_ID": ["crew_id", "Crew_ID", "crew"],
    "boq_name_clean": ["boq_name_clean", "boq_name", "BOQ_Name", "boq_code_name"],
    "IWP_ID_clean": ["iwp_id_clean", "iwp_id", "IWP_ID", "IWP_ID_clean"],
    "system_label_clean": ["system_label_clean", "system_label", "System_Label"],
    "Unit_of_Measure": ["unit_of_measure", "Unit_of_Measure", "unit", "uom"],
    "Quantity_Today": ["quantity_today", "Quantity_Today", "qty_today"],
    "Unit_Rate": ["unit_rate", "Unit_Rate"],
    "EV_DAY_VALUE": ["ev_day_value", "EV_DAY_VALUE", "ev_value"],
    "validation_status_clean": ["validation_status_clean", "validation_status", "Validation_Status"],
}


@st.cache_data(ttl=300)
def load_daily_progress_active(limit: int = 5000) -> pd.DataFrame:
    resp = supabase.table("daily_progress_active").select("*").limit(limit).execute()
    return pd.DataFrame(resp.data or [])


def option_values(df: pd.DataFrame, column: str) -> list[str]:
    if df.empty or column not in df.columns:
        return ["Все"]
    vals = df[column].dropna().astype(str).str.strip()
    vals = vals[vals != ""].unique().tolist()
    return ["Все"] + sorted(vals)


def apply_filter(df: pd.DataFrame, column: str, value: str) -> pd.DataFrame:
    if value == "Все" or column not in df.columns:
        return df
    return df[df[column].astype(str).str.strip() == value]


def safe_num(series: pd.Series | None) -> pd.Series:
    if series is None:
        return pd.Series(dtype="float64")
    return pd.to_numeric(series, errors="coerce").fillna(0.0)


def money(value: float) -> str:
    return f"{value:,.2f}".replace(",", " ").replace(".", ",")


def qty(value: float) -> str:
    return f"{value:,.3f}".replace(",", " ").replace(".", ",")


def to_excel_date(value) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return parsed.strftime("%d.%m.%Y")


def sanitize_filename_part(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]+', "_", str(value or "").strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned)
    cleaned = cleaned.strip("_")
    return cleaned or "all"


def build_pto_export_filename(
    project: str,
    month: str,
    facility: str,
    discipline: str,
) -> str:
    month_part = sanitize_filename_part(month if month != "Все" else "all_months")
    disc_part = (
        "ALL_DISCIPLINES"
        if discipline == "Все"
        else sanitize_filename_part(discipline)
    )
    if facility != "Все":
        base_part = sanitize_filename_part(facility)
    else:
        base_part = sanitize_filename_part(project if project != "Все" else "all_projects")
    return f"DP_PTO_{base_part}_{month_part}_{disc_part}.xlsx"


def resolve_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for cand in candidates:
        if cand in df.columns:
            return cand
    return None


def non_empty_count(df: pd.DataFrame, column: str | None) -> int:
    if not column or column not in df.columns:
        return 0
    vals = df[column].dropna().astype(str).str.strip()
    return int((vals != "").sum())

def build_export_df(df: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str], list[str]]:
    mapped: dict[str, str] = {}
    missing: list[str] = []
    export_df = pd.DataFrame(index=df.index)
    for target_col in EXPORT_COLUMNS:
        source_col = resolve_column(df, EXPORT_FIELD_MAP.get(target_col, [target_col]))
        if source_col:
            mapped[target_col] = source_col
            export_df[target_col] = df[source_col]
        else:
            missing.append(target_col)
    if "Work_Date" in export_df.columns:
        export_df["Work_Date"] = export_df["Work_Date"].apply(to_excel_date)
    return export_df, mapped, missing


def _resolve_work_date(
    export_df: pd.DataFrame,
    df_source: pd.DataFrame,
    mapped: dict[str, str],
    idx,
) -> object:
    if mapped.get("Work_Date") == "work_date" and "work_date" in df_source.columns:
        raw = df_source.at[idx, "work_date"]
        if raw is not None and str(raw).strip():
            parsed = pd.to_datetime(raw, errors="coerce")
            if pd.notna(parsed):
                return parsed.to_pydatetime()
    for col in WORK_DATE_FALLBACK_COLUMNS:
        if col in df_source.columns:
            raw = df_source.at[idx, col]
            if raw is not None and str(raw).strip():
                parsed = pd.to_datetime(raw, errors="coerce")
                if pd.notna(parsed):
                    return parsed.to_pydatetime()
    if "Work_Date" in export_df.columns:
        parsed = pd.to_datetime(export_df.at[idx, "Work_Date"], dayfirst=True, errors="coerce")
        if pd.notna(parsed):
            return parsed.to_pydatetime()
    return None


def _xlsx_cell_value(
    col: str,
    export_df: pd.DataFrame,
    df_source: pd.DataFrame,
    mapped: dict[str, str],
    idx,
) -> object:
    if col in XLSX_NUMERIC_COLUMNS:
        val = pd.to_numeric(export_df.at[idx, col], errors="coerce")
        return None if pd.isna(val) else float(val)
    if col in XLSX_DATE_COLUMNS:
        return _resolve_work_date(export_df, df_source, mapped, idx)
    val = export_df.at[idx, col]
    if pd.isna(val):
        return ""
    return str(val)


def _unique_nonempty_count(export_df: pd.DataFrame, column: str) -> int:
    if column not in export_df.columns:
        return 0
    vals = export_df[column].astype(str).str.strip().replace("", pd.NA).dropna()
    return int(vals.nunique())


def _export_work_date_range(
    export_df: pd.DataFrame,
    df_source: pd.DataFrame,
    mapped: dict[str, str],
) -> tuple[str, str]:
    dates: list[pd.Timestamp] = []
    for idx in export_df.index:
        dt = _resolve_work_date(export_df, df_source, mapped, idx)
        if dt is not None:
            parsed = pd.to_datetime(dt, errors="coerce")
            if pd.notna(parsed):
                dates.append(parsed)
    if not dates:
        return "—", "—"
    min_d = min(dates).strftime("%d.%m.%Y")
    max_d = max(dates).strftime("%d.%m.%Y")
    return min_d, max_d


def _export_snapshot_stats(
    export_df: pd.DataFrame,
    df_source: pd.DataFrame,
    mapped: dict[str, str],
) -> dict[str, object]:
    rows_count = len(export_df)
    qty_series = (
        safe_num(export_df["Quantity_Today"])
        if "Quantity_Today" in export_df.columns
        else pd.Series(dtype="float64")
    )
    ev_series = (
        safe_num(export_df["EV_DAY_VALUE"])
        if "EV_DAY_VALUE" in export_df.columns
        else pd.Series(dtype="float64")
    )
    qty_total = float(qty_series.sum()) if rows_count else 0.0
    ev_total = float(ev_series.sum()) if rows_count else 0.0
    zero_rows = int((qty_series.fillna(0.0) == 0.0).sum()) if rows_count else 0
    avg_ev = (ev_total / rows_count) if rows_count else 0.0
    min_d, max_d = _export_work_date_range(export_df, df_source, mapped)
    return {
        "rows": rows_count,
        "foreman": _unique_nonempty_count(export_df, "Foreman"),
        "crew": _unique_nonempty_count(export_df, "Crew_ID"),
        "iwp": _unique_nonempty_count(export_df, "IWP_ID_clean"),
        "systems": _unique_nonempty_count(export_df, "system_label_clean"),
        "qty_total": qty_total,
        "ev_total": ev_total,
        "zero_rows": zero_rows,
        "avg_ev": avg_ev,
        "period_min": min_d,
        "period_max": max_d,
    }


def _xlsx_display_label(col: str) -> str:
    labels = {
        "Work_Date": "Work_Date",
        "Month_Key": "Month_Key",
        "Facility_Building": "Facility_Building",
        "Construction_Discipline": "Construction_Discipline",
        "Foreman": "Foreman",
        "Crew_ID": "Crew_ID",
        "boq_name_clean": "boq_name_clean",
        "IWP_ID_clean": "IWP_ID_clean",
        "system_label_clean": "system_label_clean",
        "Unit_of_Measure": "Unit_of_Measure",
        "Quantity_Today": "Quantity_Today",
        "Unit_Rate": "Unit_Rate",
        "EV_DAY_VALUE": "EV_DAY_VALUE",
        "validation_status_clean": "validation_status_clean",
    }
    return labels.get(col, col)


def build_pto_xlsx_bytes(
    export_df: pd.DataFrame,
    df_source: pd.DataFrame,
    mapped: dict[str, str],
    *,
    project: str,
    month: str,
    facility: str,
    discipline: str,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    cols = [c for c in EXPORT_COLUMNS if c in export_df.columns]
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Progress PTO"

    title_fill = PatternFill("solid", fgColor=XLSX_TITLE_FILL)
    header_fill = PatternFill("solid", fgColor=XLSX_HEADER_FILL)
    zebra_fill = PatternFill("solid", fgColor=XLSX_ZEBRA_FILL)
    zero_qty_fill = PatternFill("solid", fgColor=XLSX_ZERO_QTY_FILL)
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    meta_label_font = Font(name="Calibri", size=10, bold=True, color="404040")
    meta_value_font = Font(name="Calibri", size=10, color="404040")
    summary_dark_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    total_font = Font(name="Calibri", size=10, bold=True, color="1F3864")
    note_font = Font(name="Calibri", size=8, italic=True, color="808080")
    brand_font = Font(name="Calibri", size=8, italic=True, color="808080")
    total_fill = PatternFill("solid", fgColor=XLSX_TOTAL_FILL)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="top", wrap_text=True)

    snap = _export_snapshot_stats(export_df, df_source, mapped)
    span_end = max(len(cols), 9)
    span_letter = get_column_letter(span_end)
    snap_col = XLSX_SNAPSHOT_START_COL

    ws.merge_cells(f"A1:{span_letter}1")
    title_cell = ws["A1"]
    title_cell.value = "DAILY PROGRESS EXPORT – ПТО"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 26

    export_ts = datetime.now(timezone.utc).astimezone().strftime("%d.%m.%Y %H:%M")
    filter_meta = [
        ("Проект:", project if project != "Все" else "Все проекты"),
        ("Месяц:", month if month != "Все" else "Все месяцы"),
        ("Титул:", facility if facility != "Все" else "Все титулы"),
        ("Дисциплина:", discipline if discipline != "Все" else "Все дисциплины"),
        ("Дата выгрузки:", export_ts),
    ]
    for offset, (label, value) in enumerate(filter_meta):
        row_idx = 2 + offset
        ws.cell(row=row_idx, column=1, value=label).font = meta_label_font
        val_cell = ws.cell(row=row_idx, column=2, value=value)
        val_cell.font = meta_value_font
        val_cell.alignment = left_wrap

    prod_headers = ["Строк", "Мастеров", "Звеньев", "IWP", "Систем"]
    prod_values = [snap["rows"], snap["foreman"], snap["crew"], snap["iwp"], snap["systems"]]
    comm_headers = ["Quantity", "EV", "Нулевых строк", "Средний EV/строка"]
    comm_values = [snap["qty_total"], snap["ev_total"], snap["zero_rows"], snap["avg_ev"]]

    prod_label = ws.cell(row=2, column=snap_col, value="Production")
    prod_label.font = summary_dark_font
    prod_label.fill = header_fill
    prod_label.alignment = center
    for idx, header in enumerate(prod_headers, start=1):
        cell = ws.cell(row=2, column=snap_col + idx, value=header)
        cell.font = summary_dark_font
        cell.fill = header_fill
        cell.alignment = center
    for idx, value in enumerate(prod_values, start=1):
        cell = ws.cell(row=3, column=snap_col + idx, value=value)
        cell.font = meta_value_font
        cell.alignment = center

    comm_label = ws.cell(row=4, column=snap_col, value="Commercial")
    comm_label.font = summary_dark_font
    comm_label.fill = header_fill
    comm_label.alignment = center
    for idx, header in enumerate(comm_headers, start=1):
        cell = ws.cell(row=4, column=snap_col + idx, value=header)
        cell.font = summary_dark_font
        cell.fill = header_fill
        cell.alignment = center
    for idx, value in enumerate(comm_values, start=1):
        cell = ws.cell(row=5, column=snap_col + idx, value=value)
        cell.font = meta_value_font
        cell.alignment = center
        if idx == 1:
            cell.number_format = "0.00"
        elif idx == 2:
            cell.number_format = "#,##0.00"
        elif idx == 4:
            cell.number_format = "#,##0.00"

    coverage_label = ws.cell(row=6, column=snap_col, value="Coverage")
    coverage_label.font = summary_dark_font
    coverage_label.fill = header_fill
    coverage_label.alignment = center
    period_cell = ws.cell(
        row=6,
        column=snap_col + 1,
        value=f"{snap['period_min']} → {snap['period_max']}",
    )
    period_cell.font = meta_value_font
    period_cell.alignment = left_wrap

    brand_col = max(snap_col + len(comm_headers) + 1, span_end - 1)
    brand_cell = ws.cell(row=2, column=brand_col, value="Generated by Execution OS")
    brand_cell.font = brand_font
    brand_cell.alignment = right_align
    ws.cell(row=3, column=brand_col, value="Source: daily_progress_active").font = brand_font
    ws.cell(row=3, column=brand_col).alignment = right_align

    visible_total_row = XLSX_VISIBLE_TOTAL_ROW
    data_header_row = XLSX_DATA_HEADER_ROW
    data_start_row = XLSX_DATA_START_ROW

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=data_header_row, column=col_idx, value=_xlsx_display_label(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    qty_col_idx = cols.index("Quantity_Today") + 1 if "Quantity_Today" in cols else None
    ev_col_idx = cols.index("EV_DAY_VALUE") + 1 if "EV_DAY_VALUE" in cols else None

    for data_offset, idx in enumerate(export_df.index):
        row_idx = data_start_row + data_offset
        qty_val = None
        if qty_col_idx:
            qty_val = pd.to_numeric(export_df.at[idx, "Quantity_Today"], errors="coerce")
        is_zero_qty = qty_val is not None and pd.notna(qty_val) and float(qty_val) == 0.0
        is_zebra = data_offset % 2 == 1

        for col_idx, col_name in enumerate(cols, start=1):
            value = _xlsx_cell_value(col_name, export_df, df_source, mapped, idx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left_wrap
            if col_name in XLSX_NUMERIC_COLUMNS and value is not None:
                cell.number_format = XLSX_NUM_FORMATS.get(col_name, "0.00")
            elif col_name in XLSX_DATE_COLUMNS and value is not None:
                cell.number_format = "DD.MM.YYYY"
            if is_zero_qty:
                cell.fill = zero_qty_fill
            elif is_zebra:
                cell.fill = zebra_fill

    data_last_row = data_start_row + len(export_df) - 1 if len(export_df) else data_start_row
    totals_row = data_last_row + 1
    note_row = totals_row + 1

    label_col = 1

    if len(export_df) > 0:
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=visible_total_row, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = left_wrap
        ws.cell(row=visible_total_row, column=label_col, value="ИТОГО ПО ВИДИМОМУ ФИЛЬТРУ EXCEL")
        caption_cell = ws.cell(
            row=visible_total_row,
            column=2,
            value="Пересчитывается при фильтрации внутри Excel",
        )
        caption_cell.font = note_font
        if qty_col_idx and data_last_row >= data_start_row:
            qty_letter = get_column_letter(qty_col_idx)
            ws.cell(row=visible_total_row, column=qty_col_idx).value = (
                f"=SUBTOTAL(109,{qty_letter}{data_start_row}:{qty_letter}{data_last_row})"
            )
            ws.cell(row=visible_total_row, column=qty_col_idx).number_format = "0.00"
        if ev_col_idx and data_last_row >= data_start_row:
            ev_letter = get_column_letter(ev_col_idx)
            ws.cell(row=visible_total_row, column=ev_col_idx).value = (
                f"=SUBTOTAL(109,{ev_letter}{data_start_row}:{ev_letter}{data_last_row})"
            )
            ws.cell(row=visible_total_row, column=ev_col_idx).number_format = "#,##0.00"

        ws.cell(row=totals_row, column=label_col, value="ИТОГО")
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=totals_row, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = left_wrap
        if qty_col_idx and data_last_row >= data_start_row:
            qty_letter = get_column_letter(qty_col_idx)
            ws.cell(row=totals_row, column=qty_col_idx).value = (
                f"=SUBTOTAL(109,{qty_letter}{data_start_row}:{qty_letter}{data_last_row})"
            )
            ws.cell(row=totals_row, column=qty_col_idx).number_format = "0.00"
        if ev_col_idx and data_last_row >= data_start_row:
            ev_letter = get_column_letter(ev_col_idx)
            ws.cell(row=totals_row, column=ev_col_idx).value = (
                f"=SUBTOTAL(109,{ev_letter}{data_start_row}:{ev_letter}{data_last_row})"
            )
            ws.cell(row=totals_row, column=ev_col_idx).number_format = "#,##0.00"

        note_end_col = get_column_letter(min(len(cols), 8))
        ws.merge_cells(f"A{note_row}:{note_end_col}{note_row}")
        note_cell = ws.cell(row=note_row, column=1)
        note_cell.value = (
            "ИТОГО пересчитывается при фильтрации в Excel. "
            "Верхняя сводка зафиксирована на момент выгрузки."
        )
        note_cell.font = note_font
        note_cell.alignment = left_wrap

    # Заморозка ниже header: видны snapshot, верхний ИТОГО и заголовки колонок.
    ws.freeze_panes = ws.cell(row=data_start_row, column=1).coordinate
    if cols and len(export_df) > 0:
        last_col = get_column_letter(len(cols))
        ws.auto_filter.ref = f"A{data_header_row}:{last_col}{data_last_row}"

    for col_idx, col_name in enumerate(cols, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(_xlsx_display_label(col_name))
        for row in range(data_start_row, data_start_row + len(export_df)):
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        width_cap = 70 if col_name == "boq_name_clean" else 55
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), width_cap)

    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 12, 18)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 12, 28)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


df_raw = load_daily_progress_active()

project_col = resolve_column(df_raw, ["project_code"])
month_col = resolve_column(df_raw, ["Month_Key", "month_key", "Month", "month"])
facility_col = resolve_column(df_raw, ["Facility_Building", "facility_building", "title", "Title"])
discipline_col = resolve_column(
    df_raw,
    ["Construction_Discipline", "construction_discipline", "discipline", "Discipline"],
)

if df_raw.empty:
    st.error("daily_progress_active вернула 0 строк. Проверьте Supabase/RLS/источник.")
    st.stop()

f1, f2, f3, f4 = st.columns(4)
with f1:
    if project_col:
        project = st.selectbox("Проект", option_values(df_raw, project_col))
    else:
        project = "Все"

df_project = apply_filter(df_raw, project_col or "", project)

with f2:
    if month_col:
        month = st.selectbox("Месяц", option_values(df_project, month_col))
    else:
        month = "Все"

df_month = apply_filter(df_project, month_col or "", month)

with f3:
    if facility_col:
        facility = st.selectbox("Титул / объект", option_values(df_month, facility_col))
    else:
        facility = "Все"

df_facility = apply_filter(df_month, facility_col or "", facility)

with f4:
    if discipline_col:
        discipline = st.selectbox(
            "Дисциплина",
            option_values(df_facility, discipline_col),
        )
    else:
        discipline = "Все"

df_filtered = apply_filter(df_facility, discipline_col or "", discipline).copy()

df_export, found_export_map, _missing_export_columns = build_export_df(df_filtered)

rows_count = len(df_export)
foreman_count = (
    df_export["Foreman"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
    if "Foreman" in df_export.columns
    else 0
)
crew_count = (
    df_export["Crew_ID"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
    if "Crew_ID" in df_export.columns
    else 0
)
qty_total = safe_num(df_export["Quantity_Today"] if "Quantity_Today" in df_export.columns else None).sum()
ev_total = safe_num(df_export["EV_DAY_VALUE"] if "EV_DAY_VALUE" in df_export.columns else None).sum()

k1, k2, k3, k4, k5 = st.columns(5)
k1.metric("Количество строк", rows_count)
k2.metric("Количество мастеров", int(foreman_count))
k3.metric("Количество звеньев", int(crew_count))
k4.metric("Общий объём Quantity_Today", qty(float(qty_total)))
k5.metric("Общая стоимость EV_DAY_VALUE", money(float(ev_total)))

if df_export.empty:
    st.warning("По выбранным фильтрам данные не найдены.")
    st.stop()

st.dataframe(
    df_export,
    use_container_width=True,
    hide_index=True,
    height=560,
    column_order=[col for col in EXPORT_COLUMNS if col in df_export.columns],
)

filename_xlsx = build_pto_export_filename(project, month, facility, discipline)
filename_csv = filename_xlsx.replace(".xlsx", ".csv")

try:
    xlsx_bytes = build_pto_xlsx_bytes(
        df_export,
        df_filtered,
        found_export_map,
        project=project,
        month=month,
        facility=facility,
        discipline=discipline,
    )
    st.download_button(
        "📥 Скачать Excel для ПТО (.xlsx)",
        data=xlsx_bytes,
        file_name=filename_xlsx,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False,
        type="primary",
    )
except ImportError:
    st.error("Для выгрузки Excel установите пакет: pip install openpyxl")
except Exception as exc:
    st.error(f"Не удалось сформировать Excel: {exc}")

csv_text = df_export.to_csv(index=False, sep=";")
csv_bytes = csv_text.encode("utf-8-sig")

st.download_button(
    "Скачать CSV для Excel (legacy)",
    data=csv_bytes,
    file_name=filename_csv,
    mime="text/csv",
    use_container_width=False,
)

with st.expander("Как правильно открыть CSV в Excel", expanded=False):
    st.markdown(
        """
**Рекомендуемый способ:**

1. Откройте Excel (не открывайте CSV двойным кликом)
2. Создайте новую книгу Excel
3. Перейдите: **Данные → Из текста/CSV**
4. Выберите скачанный CSV файл
5. Проверьте настройки:
   - **Кодировка:** UTF-8
   - **Разделитель:** `;` (точка с запятой)
6. Нажмите **Загрузить**

**Важно:** не открывайте CSV двойным кликом — Excel может некорректно распознать русские символы, разделители и числовые значения.
        """
    )
    st.caption(
        "Рекомендуется использовать Excel (.xlsx) — он открывается без дополнительных настроек."
    )
