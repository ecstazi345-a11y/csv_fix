"""
Unit tests for Page 24 Excel export helpers.

No product DB writes. Uses openpyxl from project venv.
"""

from __future__ import annotations

import importlib.util
import sys
import unittest
from datetime import date
from io import BytesIO
from pathlib import Path
from types import ModuleType
from typing import Any
from unittest.mock import MagicMock, patch

import pandas as pd
from openpyxl import load_workbook


PAGE_PATH = (
    Path(__file__).resolve().parents[1]
    / "pages"
    / "24_Реестр_ограничений_допуска_месячного_плана.py"
)


class _SessionState(dict):
    def __getattr__(self, name: str) -> Any:
        try:
            return self[name]
        except KeyError as exc:
            raise AttributeError(name) from exc

    def __setattr__(self, name: str, value: Any) -> None:
        self[name] = value


def _cache_data_stub(*args: Any, **kwargs: Any) -> Any:
    def decorator(fn: Any) -> Any:
        fn.clear = lambda: None  # type: ignore[attr-defined]
        return fn

    if args and callable(args[0]) and not kwargs:
        return decorator(args[0])
    return decorator


def _install_streamlit_stub() -> ModuleType:
    st = ModuleType("streamlit")
    st.session_state = _SessionState()
    st.set_page_config = lambda **kwargs: None
    st.cache_data = _cache_data_stub
    st.cache_resource = _cache_data_stub
    st.markdown = lambda *a, **k: None
    st.caption = lambda *a, **k: None
    st.info = lambda *a, **k: None
    st.warning = lambda *a, **k: None
    st.error = lambda *a, **k: None
    st.success = lambda *a, **k: None
    st.button = lambda *a, **k: False
    st.radio = lambda *a, **k: "По текущим фильтрам"
    st.text_area = lambda *a, **k: None
    st.text_input = lambda *a, **k: None
    st.selectbox = lambda *a, **k: None
    st.date_input = lambda *a, **k: date.today()
    st.columns = lambda n: [MagicMock() for _ in range(n if isinstance(n, int) else 3)]
    st.rerun = lambda: None
    st.expander = MagicMock()
    st.form = MagicMock()
    st.form_submit_button = lambda *a, **k: False
    st.dataframe = lambda *a, **k: MagicMock()
    st.download_button = lambda *a, **k: False
    st.spinner = MagicMock()
    st.divider = lambda: None
    st.column_config = MagicMock()
    st.column_config.NumberColumn = lambda *a, **k: None
    st.column_config.TextColumn = lambda *a, **k: None
    st.title = lambda *a, **k: None
    st.metric = lambda *a, **k: None
    st.write = lambda *a, **k: None
    st.text = lambda *a, **k: None
    st.checkbox = lambda *a, **k: False
    st.stop = lambda: None
    sys.modules["streamlit"] = st
    return st


def _load_page() -> ModuleType:
    for key in list(sys.modules):
        if key == "streamlit" or key.startswith("page24_excel_export_test"):
            del sys.modules[key]
    st = _install_streamlit_stub()

    class _Stop(Exception):
        pass

    st.stop = lambda: (_ for _ in ()).throw(_Stop())
    with patch(
        "services.monthly_plan_constraint_registry_service.load_constraint_registry",
        return_value=pd.DataFrame(),
    ):
        spec = importlib.util.spec_from_file_location(
            "page24_excel_export_test", PAGE_PATH
        )
        assert spec and spec.loader
        mod = importlib.util.module_from_spec(spec)
        sys.modules["page24_excel_export_test"] = mod
        try:
            spec.loader.exec_module(mod)
        except _Stop:
            pass
    return mod


class Page24ExcelExportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.mod = _load_page()

    def _sample_df(self) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "constraint_id": "aaaaaaaa-bbbb-cccc-dddd-eeeeeeee0001",
                    "project_code": "PRJ_001_БХК",
                    "month_key": "август-2026",
                    "queue": "1 очередь",
                    "facility_building": "16160-13",
                    "construction_discipline": "СМР",
                    "boq_code": "BOQ-1",
                    "boq_name": "Работа 1",
                    "work_package": "IWP-1",
                    "iwp": "IWP-1",
                    "system": "SYS-1",
                    "unit": "м",
                    "planned_qty": 10.5,
                    "plan_value": 1000.0,
                    "responsible_department": "ПТО",
                    "check_status": "HOLD",
                    "resolution_status": "OPEN",
                    "constraint_category": "Материалы",
                    "constraint_priority": "HIGH",
                    "constraint_occurred_at": "2026-08-01",
                    "constraint_created_at": "2026-08-02",
                    "days_open_real": 5,
                    "days_open": 5,
                    "problem_summary": "Нет трубы",
                    "problem_impact": "Стоп",
                    "required_action": "Поставить",
                    "problem_owner": "Проектный институт",
                    "owner_name": "Иванов",
                    "subcontractor_coordinator": "Петров",
                    "created_by": "Сидоров",
                    "effective_deadline_status": "OVERDUE",
                    "deadline_status": "OVERDUE",
                    "deadline_source": "PLAN",
                    "target_resolution_date": "2026-08-10",
                    "next_control_date": "2026-08-05",
                    "actual_resolution_date": None,
                    "value_at_risk": 1000.0,
                    "comment": "Комментарий",
                    "updated_by": "Сидоров",
                    "resolved_by": "",
                    "last_action_at": "2026-08-03",
                    "_sort_block": 0,
                    "session_tmp": "x",
                },
                {
                    "constraint_id": "aaaaaaaa-bbbb-cccc-dddd-eeeeeeee0002",
                    "project_code": "PRJ_001_БХК",
                    "month_key": "июль-2026",
                    "queue": "2 очередь",
                    "facility_building": "26160-17",
                    "construction_discipline": "ЭМ",
                    "boq_code": "BOQ-2",
                    "boq_name": "Работа 2",
                    "work_package": "IWP-2",
                    "iwp": "IWP-2",
                    "system": "SYS-2",
                    "unit": "шт",
                    "planned_qty": 2.0,
                    "plan_value": 500.0,
                    "responsible_department": "ПТО",
                    "check_status": "PASS",
                    "resolution_status": "RESOLVED",
                    "constraint_category": "Документы",
                    "constraint_priority": "LOW",
                    "constraint_occurred_at": "2026-07-01",
                    "constraint_created_at": "2026-07-02",
                    "days_open_real": 0,
                    "days_open": 0,
                    "problem_summary": "Закрыто",
                    "problem_impact": "",
                    "required_action": "",
                    "problem_owner": "Генподрядчик",
                    "owner_name": "Козлов",
                    "subcontractor_coordinator": "",
                    "created_by": "Сидоров",
                    "effective_deadline_status": "DONE",
                    "deadline_status": "DONE",
                    "deadline_source": "",
                    "target_resolution_date": "2026-07-10",
                    "next_control_date": None,
                    "actual_resolution_date": "2026-07-09",
                    "value_at_risk": 0.0,
                    "comment": "",
                    "updated_by": "Сидоров",
                    "resolved_by": "Сидоров",
                    "last_action_at": "2026-07-09",
                },
                {
                    "constraint_id": "cccccccc-cccc-cccc-cccc-cccccccc0001",
                    "project_code": "TEST_REG_R2",
                    "month_key": "test-r2-ui-2026",
                    "queue": "Не определено",
                    "facility_building": "T",
                    "construction_discipline": "СМР",
                    "boq_code": "BOQ-T",
                    "boq_name": "Test",
                    "work_package": "",
                    "iwp": "",
                    "system": "",
                    "unit": "шт",
                    "planned_qty": 1.0,
                    "plan_value": 1.0,
                    "responsible_department": "ПТО",
                    "check_status": "HOLD",
                    "resolution_status": "OPEN",
                    "constraint_category": "Test",
                    "constraint_priority": "LOW",
                    "constraint_occurred_at": "2026-08-01",
                    "constraint_created_at": "2026-08-01",
                    "days_open_real": 1,
                    "days_open": 1,
                    "problem_summary": "test",
                    "problem_impact": "",
                    "required_action": "",
                    "problem_owner": "",
                    "owner_name": "",
                    "subcontractor_coordinator": "",
                    "created_by": "",
                    "effective_deadline_status": "ON_TRACK",
                    "deadline_status": "ON_TRACK",
                    "deadline_source": "",
                    "target_resolution_date": None,
                    "next_control_date": None,
                    "actual_resolution_date": None,
                    "value_at_risk": 1.0,
                    "comment": "",
                    "updated_by": "",
                    "resolved_by": "",
                    "last_action_at": None,
                },
            ]
        )

    def test_t1_filtered_excel_created(self) -> None:
        filtered = self._sample_df().iloc[[0]].copy()
        export_df = self.mod.build_export_dataframe(filtered)
        params = self.mod.build_export_params_rows(
            formed_at=date(2026, 8, 8),
            project="PRJ_001_БХК",
            scope=self.mod.EXPORT_SCOPE_FILTERED,
            row_count=len(export_df),
            month="август-2026",
        )
        raw = self.mod.build_registry_excel_bytes(export_df, params)
        self.assertTrue(raw.startswith(b"PK"))
        self.assertGreater(len(raw), 100)

    def test_t2_row_count_matches(self) -> None:
        filtered = self._sample_df().iloc[[0, 1]].copy()
        export_df = self.mod.build_export_dataframe(filtered)
        self.assertEqual(len(export_df), 2)
        params = self.mod.build_export_params_rows(
            formed_at=date(2026, 8, 8),
            project="PRJ_001_БХК",
            scope=self.mod.EXPORT_SCOPE_FILTERED,
            row_count=len(export_df),
        )
        wb = load_workbook(BytesIO(self.mod.build_registry_excel_bytes(export_df, params)))
        ws = wb[self.mod.EXPORT_SHEET_REGISTRY]
        # title/KPI block + header + 2 data rows
        expected_last = self.mod.EXPORT_ROW_DATA_START + 1
        self.assertEqual(ws.max_row, expected_last)
        data_count = expected_last - self.mod.EXPORT_ROW_TABLE_HEADER
        self.assertEqual(data_count, 2)

    def test_t3_sheet_registry(self) -> None:
        export_df = self.mod.build_export_dataframe(self._sample_df().iloc[[0]])
        raw = self.mod.build_registry_excel_bytes(
            export_df,
            self.mod.build_export_params_rows(
                formed_at=date(2026, 8, 8),
                project="P",
                scope=self.mod.EXPORT_SCOPE_FILTERED,
                row_count=1,
            ),
        )
        wb = load_workbook(BytesIO(raw))
        self.assertIn(self.mod.EXPORT_SHEET_REGISTRY, wb.sheetnames)

    def test_t4_sheet_params(self) -> None:
        export_df = self.mod.build_export_dataframe(self._sample_df().iloc[[0]])
        raw = self.mod.build_registry_excel_bytes(
            export_df,
            self.mod.build_export_params_rows(
                formed_at=date(2026, 8, 8),
                project="P",
                scope=self.mod.EXPORT_SCOPE_FILTERED,
                row_count=1,
            ),
        )
        wb = load_workbook(BytesIO(raw))
        self.assertIn(self.mod.EXPORT_SHEET_PARAMS, wb.sheetnames)

    def test_t5_technical_columns_hidden(self) -> None:
        export_df = self.mod.build_export_dataframe(self._sample_df().iloc[[0]])
        self.assertNotIn("_sort_block", export_df.columns)
        self.assertNotIn("session_tmp", export_df.columns)
        self.assertIn("Проект", export_df.columns)
        self.assertIn("ID ограничения", export_df.columns)
        self.assertEqual(list(export_df.columns)[-1], "ID ограничения")

    def test_t6_full_project_excludes_test(self) -> None:
        full = self._sample_df()
        filtered = full.iloc[[0]].copy()
        out = self.mod.resolve_export_source_df(
            full_df=full,
            filtered_df=filtered,
            scope=self.mod.EXPORT_SCOPE_FULL_PROJECT,
            project="PRJ_001_БХК",
        )
        self.assertEqual(len(out), 2)
        self.assertTrue((out["project_code"] == "PRJ_001_БХК").all())
        self.assertNotIn("TEST_REG_R2", set(out["project_code"].astype(str)))

    def test_t7_empty_safe(self) -> None:
        empty = pd.DataFrame()
        export_df = self.mod.build_export_dataframe(empty)
        self.assertEqual(len(export_df), 0)
        self.assertIn("Проект", export_df.columns)
        source = self.mod.resolve_export_source_df(
            full_df=empty,
            filtered_df=empty,
            scope=self.mod.EXPORT_SCOPE_FILTERED,
            project="PRJ_001_БХК",
        )
        self.assertTrue(source.empty)
        # Empty workbook still builds (UI won't offer download).
        raw = self.mod.build_registry_excel_bytes(
            export_df,
            self.mod.build_export_params_rows(
                formed_at=date(2026, 8, 8),
                project="P",
                scope=self.mod.EXPORT_SCOPE_FILTERED,
                row_count=0,
            ),
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb[self.mod.EXPORT_SHEET_REGISTRY]
        self.assertEqual(ws.max_row, self.mod.EXPORT_ROW_TABLE_HEADER)
        self.assertIn(
            "РЕЕСТР ОГРАНИЧЕНИЙ ПРОИЗВОДСТВА РАБОТ",
            str(ws.cell(row=self.mod.EXPORT_ROW_TITLE, column=1).value),
        )

    def test_russian_labels_and_freeze_autofilter(self) -> None:
        export_df = self.mod.build_export_dataframe(self._sample_df().iloc[[0]])
        raw = self.mod.build_registry_excel_bytes(
            export_df,
            self.mod.build_export_params_rows(
                formed_at=date(2026, 8, 8),
                project="PRJ_001_БХК",
                scope=self.mod.EXPORT_SCOPE_FILTERED,
                row_count=1,
                month="август-2026",
            ),
        )
        wb = load_workbook(BytesIO(raw))
        ws = wb[self.mod.EXPORT_SHEET_REGISTRY]
        header_row = self.mod.EXPORT_ROW_TABLE_HEADER
        data_row = self.mod.EXPORT_ROW_DATA_START
        self.assertEqual(ws.cell(row=header_row, column=1).value, "Проект")
        self.assertTrue(ws.cell(row=header_row, column=1).font.bold)
        self.assertEqual(ws.freeze_panes, f"C{data_row}")
        self.assertIsNotNone(ws.auto_filter.ref)
        self.assertTrue(
            str(ws.auto_filter.ref).startswith(f"A{header_row}:"),
            msg=ws.auto_filter.ref,
        )
        headers = [cell.value for cell in ws[header_row]]
        check_idx = headers.index("Статус проверки") + 1
        from openpyxl.utils import get_column_letter

        check_cell = ws[f"{get_column_letter(check_idx)}{data_row}"]
        self.assertEqual(check_cell.value, "Удержание")
        money_col = headers.index("Стоимость под риском") + 1
        cell = ws[f"{get_column_letter(money_col)}{data_row}"]
        self.assertEqual(cell.value, 1000.0)

    def test_formatting_contract_enterprise_report(self) -> None:
        export_df = self.mod.build_export_dataframe(self._sample_df().iloc[[0]])
        before_cols = list(export_df.columns)
        before_values = export_df.copy()
        params = self.mod.build_export_params_rows(
            formed_at=date(2026, 8, 8),
            project="PRJ_001_БХК",
            scope=self.mod.EXPORT_SCOPE_FILTERED,
            row_count=1,
            month="август-2026",
        )
        raw = self.mod.build_registry_excel_bytes(export_df, params)
        wb = load_workbook(BytesIO(raw))

        self.assertEqual(
            wb.sheetnames,
            [self.mod.EXPORT_SHEET_REGISTRY, self.mod.EXPORT_SHEET_PARAMS],
        )
        ws = wb[self.mod.EXPORT_SHEET_REGISTRY]
        title = ws.cell(row=self.mod.EXPORT_ROW_TITLE, column=1).value
        self.assertEqual(title, self.mod.EXPORT_TITLE)
        self.assertTrue(ws.merged_cells.ranges)
        merged = {str(r) for r in ws.merged_cells.ranges}
        self.assertTrue(any(m.startswith("A1:") for m in merged))

        header_row = self.mod.EXPORT_ROW_TABLE_HEADER
        header_cell = ws.cell(row=header_row, column=1)
        self.assertTrue(header_cell.font.bold)
        self.assertEqual(
            str(header_cell.font.color.rgb)[-6:].upper(),
            self.mod.COLOR_WHITE,
        )
        self.assertEqual(
            str(header_cell.fill.fgColor.rgb)[-6:].upper(),
            self.mod.COLOR_BLUE,
        )
        self.assertEqual(ws.freeze_panes, f"C{self.mod.EXPORT_ROW_DATA_START}")
        self.assertTrue(
            str(ws.auto_filter.ref).startswith(f"A{header_row}:"),
        )
        self.assertFalse(
            str(ws.auto_filter.ref).startswith("A1:"),
            msg="AutoFilter must not include title/KPI block",
        )
        headers = [cell.value for cell in ws[header_row]]
        self.assertEqual(headers, before_cols)
        self.assertEqual(headers[0], "Проект")
        self.assertIn("Статус проверки", headers)
        self.assertEqual(headers[-1], "ID ограничения")

        self.assertEqual(ws.page_setup.orientation, "landscape")
        self.assertEqual(ws.page_setup.fitToWidth, 1)

        data_row = self.mod.EXPORT_ROW_DATA_START
        for col_idx, col_name in enumerate(before_cols, start=1):
            expected = before_values.iloc[0][col_name]
            actual = ws.cell(row=data_row, column=col_idx).value
            if pd.isna(expected):
                self.assertTrue(actual is None or actual == "" or pd.isna(actual))
            elif isinstance(expected, date) and hasattr(actual, "date"):
                # openpyxl round-trips date cells as datetime
                self.assertEqual(actual.date(), expected)
            else:
                self.assertEqual(actual, expected)

        ws_params = wb[self.mod.EXPORT_SHEET_PARAMS]
        self.assertEqual(ws_params["A1"].value, "ПАРАМЕТРЫ ВЫГРУЗКИ")
        self.assertEqual(ws_params["A2"].value, "Параметр")
        self.assertEqual(ws_params["B2"].value, "Значение")
        param_keys = {
            ws_params.cell(row=r, column=1).value
            for r in range(3, ws_params.max_row + 1)
        }
        self.assertIn("Проект", param_keys)
        self.assertIn("Объём выгрузки", param_keys)
        self.assertIn("Количество строк в выгрузке", param_keys)

        # Dataframe unchanged by builder
        self.assertEqual(list(export_df.columns), before_cols)
        self.assertEqual(len(export_df), 1)


if __name__ == "__main__":
    unittest.main()
